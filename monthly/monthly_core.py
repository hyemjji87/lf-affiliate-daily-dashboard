"""
LF몰 제휴 월마감 대시보드 - 월별 집계 & 피벗 기반 파생 로직

핵심: 업로드 raw의 `당월인증`·`제휴사`·`거래액_VAT제외`·UV `제휴사명`은 원본이 '피벗 시트를
참조하는 엑셀 수식'이라, 파일에 계산 캐시가 없으면 값이 비어 보인다. 그래서 이 모듈은
엑셀 수식을 파이썬으로 그대로 재현(피벗에서 매핑)하여, 캐시 유무와 무관하게 항상 정확한
값을 만든다.

재현하는 수식 (인증거래액 시트 기준)
- 당월인증(AC)  = IF(LEFT(정산일,7)=LEFT(최초인증일,7),"Y","N")   → 정산 YYYY-MM == 최초인증 YYYY-MM
- 제휴사(AB)     = VLOOKUP(제휴처구분3, 피벗!B:C, 2)                 → 코드→약칭
- 거래액_VAT제외 = 거래액 / 1.1
- 인증회원 제휴사 = VLOOKUP(제휴처구분4, 피벗!B:C, 2)
- 유입 제휴사명   = AF코드 → 피벗 I:J(930명) → 피벗 B:C(약칭)
- 주차           = 피벗 <주차기준> 일자 → 라벨(YY_M_W)  →  마감(년,월) & 주차번호
"""
from __future__ import annotations
import re
import datetime as dt

import numpy as np
import pandas as pd

import lf_analysis as A

STATUSES = A.STATUSES  # ["신규", "WIN-BACK", "기존"]
SEG_LABEL = {"신규": "신규", "WIN-BACK": "윈백", "기존": "기존"}


# ======================================================================================
# 1) 피벗 파싱: 3개 블록(제휴사 구분 / 주차기준 / 제휴사명 정제)을 각각 dict로
# ======================================================================================
def parse_pivot(pivot_df: pd.DataFrame | None) -> dict:
    """피벗 시트(위치 기반)를 파싱해 매핑 dict 반환.
    반환: code_to_short, af_to_930, date_to_label, date_to_close(=(y,m)), date_to_week(int)
    """
    maps = {"code_to_short": {}, "af_to_930": {}, "date_to_label": {},
            "date_to_close": {}, "date_to_week": {}}
    if pivot_df is None or pivot_df.empty:
        return maps
    p = pivot_df.reset_index(drop=True)

    def col(i):
        return p.iloc[:, i] if (i is not None and 0 <= i < p.shape[1]) else pd.Series([], dtype=object)

    # 리더(calamine/openpyxl)에 따라 빈 선두열 트림 여부가 달라 열 인덱스가 흔들린다.
    # 따라서 헤더 마커(<제휴사 구분>/<주차기준>/<제휴사명 정제>)의 열 위치를 찾아 상대 참조한다.
    a_start = b_start = c_start = None
    for ri in range(min(3, len(p))):
        for j in range(p.shape[1]):
            s = str(p.iat[ri, j])
            if a_start is None and "제휴사" in s and "구분" in s:
                a_start = j
            if b_start is None and "주차" in s:
                b_start = j
            if c_start is None and "정제" in s:
                c_start = j
        if a_start is not None and b_start is not None and c_start is not None:
            break

    # 블록 A (제휴사 구분): [start]=코드(000127-SKT / 930007-현대3F), [start+1]=약칭(SKT)
    for code, short in zip(col(a_start), col(a_start + 1 if a_start is not None else None)):
        if pd.notna(code) and pd.notna(short) and str(short).strip() and not str(code).startswith("<"):
            maps["code_to_short"][str(code).strip()] = str(short).strip()

    # 블록 C (제휴사명 정제): [start]=AF코드, [start+1]=930명
    for af, n930 in zip(col(c_start), col(c_start + 1 if c_start is not None else None)):
        if pd.notna(af) and pd.notna(n930) and not str(af).startswith("<"):
            maps["af_to_930"][str(af).strip()] = str(n930).strip()

    # 블록 B (주차기준): [start]=일자, [start+2]=주차라벨(YY_M_W)
    dates = pd.to_datetime(col(b_start), errors="coerce")
    labels = col(b_start + 2 if b_start is not None else None)
    for d, lab in zip(dates, labels):
        if pd.isna(d) or pd.isna(lab) or str(lab).startswith("<"):
            continue
        key = d.date()
        lab = str(lab).strip()
        maps["date_to_label"][key] = lab
        ym, wk = _parse_week_label(lab)
        if ym:
            maps["date_to_close"][key] = ym
            maps["date_to_week"][key] = wk
    return maps


def _parse_week_label(lab: str):
    """'26_1_1' → ((2026,1), 1). 실패 시 (None, None)."""
    m = re.match(r"^\s*(\d{2,4})[_\-.](\d{1,2})[_\-.](\d{1,2})", str(lab))
    if not m:
        return None, None
    y = int(m.group(1))
    if y < 100:
        y += 2000
    return (y, int(m.group(2))), int(m.group(3))


# ======================================================================================
# 2) enrich: 파생 컬럼을 실제 값으로 채움 (수식 재현). 이미 값이 있으면 보존(전년 파일 대비).
# ======================================================================================
def enrich(data: dict, maps: dict) -> dict:
    amt, uv, cert = data.get("amt"), data.get("uv"), data.get("cert")
    c2s = maps.get("code_to_short", {})
    af930 = maps.get("af_to_930", {})
    d2close = maps.get("date_to_close", {})
    d2week = maps.get("date_to_week", {})
    d2label = maps.get("date_to_label", {})

    if amt is not None and not amt.empty:
        # 거래액_VAT제외 = 거래액/1.1 (캐시 비었으면 재계산)
        if "거래액" in amt.columns:
            gross = pd.to_numeric(amt["거래액"], errors="coerce")
            vat = pd.to_numeric(amt.get("거래액_VAT제외"), errors="coerce")
            need = vat.isna() | (vat == 0)
            amt["거래액_VAT제외"] = vat.where(~need, gross / 1.1).fillna(0.0)
        # 당월인증: 정산 YYYY-MM == 최초인증 YYYY-MM
        cur = (amt["당월인증"].astype(str).str.strip().str.upper()
               if "당월인증" in amt.columns else pd.Series("", index=amt.index))
        has = cur.isin(["Y", "N"])
        if not has.all() and "최초인증일일" in amt.columns:
            s = pd.to_datetime(amt["정산일시일"], errors="coerce").dt.strftime("%Y-%m")
            fst = pd.to_datetime(amt["최초인증일일"], errors="coerce").dt.strftime("%Y-%m")
            derived = pd.Series(np.where((s == fst) & s.notna(), "Y", "N"), index=amt.index)
            amt["당월인증"] = cur.where(has, derived)
        else:
            amt["당월인증"] = cur.where(has, "N")
        # 제휴사: 코드(제휴처구분3)→약칭
        if "제휴처구분3" in amt.columns:
            need_p = "제휴사" not in amt.columns or amt["제휴사"].isna().all()
            if need_p:
                key = amt["제휴처구분3"].astype(str).str.strip()
                amt["제휴사"] = key.map(c2s).fillna(key.str.replace(r"^[0-9]+-", "", regex=True))
        # 마감월 / 주차 (피벗 기준, 없으면 캘린더 폴백)
        dser = pd.to_datetime(amt["정산일시일"], errors="coerce").dt.date
        amt["_마감"] = dser.map(lambda d: _ym_str(d2close.get(d)) or _cal_ym(d))
        amt["_주차"] = dser.map(lambda d: d2label.get(d) or "")
        amt["_주차N"] = dser.map(lambda d: d2week.get(d) or 0)

    if cert is not None and not cert.empty:
        code_col = "제휴처구분4" if "제휴처구분4" in cert.columns else "제휴처구분3"
        if code_col in cert.columns:
            need_p = "제휴사" not in cert.columns or cert["제휴사"].isna().all()
            if need_p:
                key = cert[code_col].astype(str).str.strip()
                cert["제휴사"] = key.map(c2s).fillna(key.str.replace(r"^[0-9]+-", "", regex=True))
        dser = pd.to_datetime(cert["인증일시일"], errors="coerce").dt.date
        cert["_마감"] = dser.map(lambda d: _ym_str(d2close.get(d)) or _cal_ym(d))
        cert["_주차"] = dser.map(lambda d: d2label.get(d) or "")

    if uv is not None and not uv.empty:
        need_p = "제휴사명" not in uv.columns or uv["제휴사명"].isna().all() or (uv["제휴사명"].astype(str).str.strip() == "").all()
        if need_p and "AF코드" in uv.columns:
            af = uv["AF코드"].astype(str).str.strip()
            uv["제휴사명"] = af.map(lambda a: c2s.get(af930.get(a, ""), af930.get(a, a)))
        dser = pd.to_datetime(uv["★일자일"], errors="coerce").dt.date
        uv["_마감"] = dser.map(lambda d: _ym_str(d2close.get(d)) or _cal_ym(d))
        uv["_주차"] = dser.map(lambda d: d2label.get(d) or "")

    # 메모리 절감: 집계에 쓰지 않는 무거운 컬럼(상품명 등)을 제거한다.
    # (Streamlit Cloud 메모리 한도 대비 — raw 여러 달 업로드 시 수십만 행 × 긴 문자열)
    _prune(amt, ["정산일시일", "거래액_VAT제외", "정산구분", "고객번호", "기존/win-back/신규",
                 "당월인증", "제휴사", "BPU", "물리대카테", "Admin브랜드명",
                 "회원등급", "첫구매주문건여부", "정상이월구분", "상품코드",  # 향후 분석용 보존
                 "_마감", "_주차", "_주차N"])
    _prune(uv, ["★일자일", "★UV", "AF코드", "제휴사명", "_마감", "_주차"])
    _prune(cert, ["인증일시일", "총합계", "기존", "WIN-BACK", "신규", "제휴사", "_마감", "_주차"])
    return data


def _prune(df, keep):
    """DataFrame에서 keep 목록에 있는 컬럼만 남긴다(존재하는 것만). in-place."""
    if df is None or df.empty:
        return
    drop = [c for c in df.columns if c not in keep]
    if drop:
        df.drop(columns=drop, inplace=True)


def load_raw_pruned(list_of_bytes: list) -> dict:
    """raw 파일들을 '한 개씩' 로드→파생/프루닝→누적→마지막에 concat 한다.
    여러 달 raw를 한꺼번에 메모리에 올리지 않아 최대 메모리를 크게 낮춘다
    (Streamlit Cloud OOM 방지). 반환: {amt, uv, cert, pivot}. amt/uv/cert는 이미 enrich·프루닝됨."""
    import gc
    empty = {"amt": pd.DataFrame(), "uv": pd.DataFrame(), "cert": pd.DataFrame(), "pivot": None}
    if not list_of_bytes:
        return empty
    pivot = None
    maps = None
    amts, uvs, certs = [], [], []
    for b in list_of_bytes:
        d = A.classify_workbook(b)          # 파일 1개(전체 컬럼)
        if maps is None and d.get("pivot") is not None:
            pivot = d["pivot"]
            maps = parse_pivot(pivot)
        enrich(d, maps or {})               # 파생값 계산 후 불필요 컬럼 프루닝
        for key, acc in (("amt", amts), ("uv", uvs), ("cert", certs)):
            f = d.get(key)
            if f is not None and not f.empty:
                acc.append(f)
        del d
        gc.collect()
    out = {
        "amt": pd.concat(amts, ignore_index=True) if amts else pd.DataFrame(),
        "uv": pd.concat(uvs, ignore_index=True) if uvs else pd.DataFrame(),
        "cert": pd.concat(certs, ignore_index=True) if certs else pd.DataFrame(),
        "pivot": pivot,
    }
    del amts, uvs, certs
    gc.collect()
    return out


def _ym_str(ym):
    return f"{ym[0]}-{ym[1]:02d}" if ym else None


def _cal_ym(d):
    return f"{d.year}-{d.month:02d}" if isinstance(d, dt.date) else None


# ======================================================================================
# 3) 기간 유틸: 마감월별 날짜 목록 (월마감 / MTD)
# ======================================================================================
def available_months(*datasets) -> list:
    ms = set()
    for d in datasets:
        for df in (d.get("amt"), d.get("uv"), d.get("cert")):
            if df is not None and not df.empty and "_마감" in df.columns:
                ms |= set(df["_마감"].dropna().unique().tolist())
    return sorted(m for m in ms if m)


def month_dates(data: dict, ym: str, maps: dict, mtd_day: int | None = None) -> list:
    """해당 마감월(ym)에 속하는 실제 날짜 목록. 피벗 date_to_close 우선, 없으면 데이터에서 수집.
    mtd_day가 주어지면 캘린더 일자 <= mtd_day 만."""
    dates = set()
    for d, close in maps.get("date_to_close", {}).items():
        if _ym_str(close) == ym:
            dates.add(d)
    if not dates:  # 피벗에 없으면 데이터에서
        for df, dc in ((data.get("amt"), "정산일시일"), (data.get("uv"), "★일자일"), (data.get("cert"), "인증일시일")):
            if df is not None and not df.empty and "_마감" in df.columns:
                sub = df[df["_마감"] == ym]
                dates |= set(pd.to_datetime(sub[dc], errors="coerce").dt.date.dropna().tolist())
    out = sorted(d for d in dates if d)
    if mtd_day is not None:
        out = [d for d in out if d.day <= mtd_day]
    return out


def weeks_of_month(data: dict, ym: str) -> list:
    """해당 마감월의 주차 라벨 목록(정렬)."""
    labels = set()
    amt = data.get("amt")
    for df in (data.get("amt"), data.get("uv"), data.get("cert")):
        if df is not None and not df.empty and "_마감" in df.columns and "_주차" in df.columns:
            sub = df[df["_마감"] == ym]
            labels |= set(x for x in sub["_주차"].dropna().unique().tolist() if x)
    return sorted(labels, key=lambda l: _parse_week_label(l)[1] or 0)


def week_dates(data: dict, label: str) -> list:
    dates = set()
    for df, dc in ((data.get("amt"), "정산일시일"), (data.get("uv"), "★일자일"), (data.get("cert"), "인증일시일")):
        if df is not None and not df.empty and "_주차" in df.columns:
            sub = df[df["_주차"] == label]
            dates |= set(pd.to_datetime(sub[dc], errors="coerce").dt.date.dropna().tolist())
    return sorted(d for d in dates if d)


# ======================================================================================
# 4) 핵심 지표 묶음 (한 기간에 대한 UV/인증/당월인증거래액/전체거래액/고객수/객단가/CR)
# ======================================================================================
def metric_bundle(data: dict, dates: list, partners=None, exclude_af=None) -> dict:
    """한 기간(dates)에 대한 지표 묶음. 세그(신규/윈백/기존) 포함.
    반환 dict 키: uv, cert(전체/신규/WIN-BACK/기존), amt_all, amt_cert (각 tot/net/cust_tot/cust_net, 세그별)"""
    uv = A.uv_total(data.get("uv", pd.DataFrame()), dates, partners, exclude_af=exclude_af) if data.get("uv") is not None else 0
    cert = A.cert_summary(data.get("cert", pd.DataFrame()), dates, partners) if data.get("cert") is not None else {"전체": 0, "기존": 0, "WIN-BACK": 0, "신규": 0}
    amt_all = A.amt_summary(data.get("amt", pd.DataFrame()), dates, partners, cert_only=False)
    amt_cert = A.amt_summary(data.get("amt", pd.DataFrame()), dates, partners, cert_only=True)
    return {"uv": uv, "cert": cert, "amt_all": amt_all, "amt_cert": amt_cert}


# ======================================================================================
# 5) 기여도(상승/하락) & BPU 추가분석
# ======================================================================================
def contribution_table(amt_cur, dates_cur, amt_prev, dates_prev, groupcol, pt="net",
                       cert_only=True, top=12) -> pd.DataFrame:
    """그룹(카테고리/브랜드)별 전년비·비중·기여도. pt: 'net'(순결제)/'tot'(총결제).
    기여도 = 그룹 증감 / 전체 증감 (합계 100%). |증감| 큰 순으로 top개."""
    cur = A.group_amt_table(amt_cur, dates_cur, None, groupcol, cert_only=cert_only)
    cur = _pick(cur, pt).rename(columns={"val": "cur"})
    if amt_prev is not None and not amt_prev.empty:
        prev = A.group_amt_table(amt_prev, dates_prev, None, groupcol, cert_only=cert_only)
        prev = _pick(prev, pt).rename(columns={"val": "prev"})
        m = pd.merge(cur, prev, on=groupcol, how="outer").fillna(0.0)
    else:
        m = cur.copy()
        m["prev"] = np.nan
    m["delta"] = m["cur"] - m["prev"].fillna(0)
    tot_cur = m["cur"].sum()
    tot_delta = m["delta"].sum()
    m["share"] = np.where(tot_cur != 0, m["cur"] / tot_cur, 0)
    m["yoy"] = np.where(m["prev"].fillna(0) != 0, m["cur"] / m["prev"] - 1, np.nan)
    m["contrib"] = np.where(tot_delta != 0, m["delta"] / tot_delta, 0)
    m = m.reindex(m["delta"].abs().sort_values(ascending=False).index)
    return m.head(top).reset_index(drop=True)


def _pick(df, pt):
    if df.empty:
        return pd.DataFrame(columns=[df.columns[0] if len(df.columns) else "grp", "val"])
    gc = df.columns[0]
    return df[[gc, pt]].rename(columns={pt: "val"})


def bpu_decomposition(amt_cur, dates_cur, amt_prev, dates_prev, pt="net", cert_only=True) -> pd.DataFrame:
    """BPU별 객단가 분해: 거래액 증감을 '고객수 효과'와 '객단가 효과'로 분리.
    ΔRev = Δcust·aov_prev + cust_prev·Δaov + Δcust·Δaov (교차항은 객단가 효과에 합산)."""
    def _agg(amt, dates):
        sub = A._filter_amt(amt, dates, None, cert_only)
        if sub.empty:
            return pd.DataFrame(columns=["BPU", "rev", "cust"])
        base = sub[sub["정산구분"] == "판매"] if pt == "tot" else sub
        rev = base.groupby("BPU")["거래액_VAT제외"].sum()
        cust = base.groupby("BPU")["고객번호"].nunique()
        return pd.DataFrame({"rev": rev, "cust": cust}).reset_index()

    cur = _agg(amt_cur, dates_cur).rename(columns={"rev": "rev_c", "cust": "cust_c"})
    if amt_prev is not None and not amt_prev.empty:
        prev = _agg(amt_prev, dates_prev).rename(columns={"rev": "rev_p", "cust": "cust_p"})
        m = pd.merge(cur, prev, on="BPU", how="outer").fillna(0.0)
    else:
        m = cur.copy()
        m["rev_p"] = np.nan
        m["cust_p"] = np.nan
    m["aov_c"] = np.where(m["cust_c"] != 0, m["rev_c"] / m["cust_c"], 0)
    m["aov_p"] = np.where(m["cust_p"].fillna(0) != 0, m["rev_p"] / m["cust_p"], np.nan)
    m["d_rev"] = m["rev_c"] - m["rev_p"].fillna(0)
    m["cust_effect"] = (m["cust_c"] - m["cust_p"].fillna(0)) * m["aov_p"].fillna(0)
    m["aov_effect"] = m["d_rev"] - m["cust_effect"]
    return m.sort_values("rev_c", ascending=False).reset_index(drop=True)


def return_rate_table(amt_cur, dates_cur, groupcol="BPU", cert_only=True) -> pd.DataFrame:
    """반품률 = (총결제 - 순결제) / 총결제, 그룹별."""
    t = A.group_amt_table(amt_cur, dates_cur, None, groupcol, cert_only=cert_only)
    if t.empty:
        return t
    t["반품률"] = np.where(t["tot"] != 0, (t["tot"] - t["net"]) / t["tot"], 0)
    return t.sort_values("tot", ascending=False).reset_index(drop=True)


def new_penetration_table(amt_cur, dates_cur, groupcol="BPU", pt="net", cert_only=True) -> pd.DataFrame:
    """신규 침투율 = 신규 거래액 / 전체 거래액, 그룹별."""
    sub = A._filter_amt(amt_cur, dates_cur, None, cert_only)
    if sub.empty:
        return pd.DataFrame(columns=[groupcol, "전체", "신규", "침투율"])
    base = sub[sub["정산구분"] == "판매"] if pt == "tot" else sub
    total = base.groupby(groupcol)["거래액_VAT제외"].sum()
    new = base[base["기존/win-back/신규"] == "신규"].groupby(groupcol)["거래액_VAT제외"].sum()
    out = pd.DataFrame({"전체": total, "신규": new}).fillna(0.0).reset_index()
    out["침투율"] = np.where(out["전체"] != 0, out["신규"] / out["전체"], 0)
    return out.sort_values("전체", ascending=False).reset_index(drop=True)


def pareto_summary(amt_cur, dates_cur, groupcol="Admin브랜드명", pt="net", cert_only=True) -> dict:
    """상위 5/10/20 브랜드가 거래액에서 차지하는 누적 비중."""
    t = A.group_amt_table(amt_cur, dates_cur, None, groupcol, cert_only=cert_only)
    if t.empty:
        return {}
    vals = t[pt].sort_values(ascending=False).reset_index(drop=True)
    total = vals.sum()
    if total == 0:
        return {}
    out = {"total_n": int((vals > 0).sum())}
    for n in (5, 10, 20):
        out[f"top{n}"] = float(vals.head(n).sum() / total)
    return out


# ======================================================================================
# 6) 목표 / LF몰 전체거래액 로더 (월별 wide 템플릿)
# ======================================================================================
def _norm_ym(v) -> str | None:
    """월 헤더 문자열을 'YYYY-MM'으로 정규화. 연 2000~2099·월 1~12만 인정
    (숫자 값 '2633.0' 등을 연-월로 오인하지 않도록 범위 검증)."""
    if isinstance(v, (dt.date, dt.datetime)):
        return f"{v.year}-{v.month:02d}"
    s = str(v).strip()
    y = mth = None
    m = re.match(r"^(20\d{2})[-./](\d{1,2})(?!\d)", s)  # 2026-01 / 2026.1
    if m:
        y, mth = int(m.group(1)), int(m.group(2))
    else:
        digits = re.sub(r"[^0-9]", "", s)
        if len(digits) == 6:  # 202601 (정확히 6자리)
            y, mth = int(digits[:4]), int(digits[4:6])
    if y is not None and 2000 <= y <= 2099 and 1 <= mth <= 12:
        return f"{y}-{mth:02d}"
    return None


METRIC_KEYS = ["UV", "인증자수", "당월인증거래액", "거래액"]


def _canon_metric(s):
    s = str(s)
    if "당월인증" in s:
        return "당월인증거래액"
    if "UV" in s.upper():
        return "UV"
    if "인증" in s:  # 인증자수 / 인증회원수 등
        return "인증자수"
    if "거래액" in s or "매출" in s:
        return "거래액"
    return None


def _canon_seg(s):
    s = str(s).strip().lower()
    if s in ("신규", "new"):
        return "신규"
    if "win" in s or "윈백" in s:
        return "윈백"
    if s in ("기존", "existing"):
        return "기존"
    if s in ("전체", "합계", "계", "total"):
        return "전체"
    return None


def _find_month_header(data):
    """월(YYYY-MM) 헤더가 있는 행 index와 {열idx: ym} 매핑을 찾는다(제목행이 위에 있어도 대응)."""
    best_i, best_map = None, {}
    for i, row in enumerate(data[:20]):
        m = {}
        for j, c in enumerate(row or []):
            ym = _norm_ym(c)
            if ym:
                m[j] = ym
        if len(m) > len(best_map):
            best_i, best_map = i, m
    return best_i, best_map


def load_target(file_bytes: bytes) -> pd.DataFrame:
    """목표 파일 로더 — 두 가지 레이아웃을 모두 인식한다.
    (A) 구조형: 열 [지표, 세그, <월들…>]
    (B) 섹션형(사용자 양식): 월=열, 지표명 헤더행 아래 신규/윈백/기존 행이 이어짐.
    지표 ∈ {UV, 인증자수, 당월인증거래액, 거래액}. 반환(long): 지표, 세그, ym, 목표.
    '전체'가 없으면 세그 합으로 자동 생성."""
    sheets = A._read_all_sheets(file_bytes)
    best = None
    for d in sheets.values():
        if not d:
            continue
        hi, mm = _find_month_header(d)
        if hi is not None and len(mm) >= 3 and (best is None or len(mm) > len(best[2])):
            best = (d, hi, mm)
    if best is None:
        return pd.DataFrame(columns=["지표", "세그", "ym", "목표"])
    data, hi, ym_cols = best
    header = [str(c).strip() for c in data[hi]]
    i_metric = next((j for j, c in enumerate(header) if "지표" in c), None)
    i_seg = next((j for j, c in enumerate(header) if "세그" in c or "구분" in c), None)

    rows = []

    def _add(metric, seg, ym, v):
        if metric is None or seg is None or v in (None, ""):
            return
        try:
            rows.append({"지표": metric, "세그": seg, "ym": ym, "목표": float(v)})
        except (TypeError, ValueError):
            pass

    if i_metric is not None:  # (A) 구조형
        if i_seg is None:
            i_seg = i_metric + 1
        for r in data[hi + 1:]:
            if not r or i_metric >= len(r) or r[i_metric] in (None, ""):
                continue
            metric = _canon_metric(r[i_metric])
            seg = _canon_seg(r[i_seg]) if (i_seg < len(r) and r[i_seg] not in (None, "")) else "전체"
            for j, ym in ym_cols.items():
                if j < len(r):
                    _add(metric, seg or "전체", ym, r[j])
    else:  # (B) 섹션형
        label_cols = [j for j in range(len(header)) if j not in ym_cols]
        lc = label_cols[0] if label_cols else 0
        current = None
        for r in data[hi + 1:]:
            if not r:
                continue
            label = str(r[lc]).strip() if (lc < len(r) and r[lc] not in (None, "")) else ""
            if not label:
                continue
            seg = _canon_seg(label)
            if seg:  # 세그 행
                for j, ym in ym_cols.items():
                    if j < len(r):
                        _add(current, seg, ym, r[j])
            else:
                m = _canon_metric(label)
                if m:
                    current = m
                    # 지표 헤더행이 값을 직접 담고 있으면(UV처럼 세그 없는 지표) 전체로 적재
                    if any(j < len(r) and r[j] not in (None, "") for j in ym_cols):
                        for j, ym in ym_cols.items():
                            if j < len(r):
                                _add(m, "전체", ym, r[j])

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    out = [df]
    for (metric, ym), g in df.groupby(["지표", "ym"]):
        if "전체" not in set(g["세그"]):
            segs = g[g["세그"].isin(["신규", "윈백", "기존"])]
            if not segs.empty:
                out.append(pd.DataFrame([{"지표": metric, "세그": "전체", "ym": ym, "목표": segs["목표"].sum()}]))
    return pd.concat(out, ignore_index=True)


def target_value(target_df, 지표, ym, 세그="전체"):
    if target_df is None or target_df.empty:
        return None
    sub = target_df[(target_df["지표"] == 지표) & (target_df["ym"] == ym) & (target_df["세그"] == 세그)]
    return float(sub["목표"].iloc[0]) if not sub.empty else None


def load_lfmall(file_bytes: bytes) -> pd.DataFrame:
    """LF몰 전체거래액 로더. 기대 구조: 열 [구분, <월들...>], 구분 ∈ {목표, 실제}.
    (또는 열 [월, 목표, 실제]). 반환: ym, 목표, 실제."""
    sheets = A._read_all_sheets(file_bytes)
    # 월 헤더가 가장 많은 시트/행을 선택(제목행이 있어도 대응)
    best = None
    for d in sheets.values():
        if not d:
            continue
        hi, mm = _find_month_header(d)
        if hi is not None and len(mm) >= 3 and (best is None or len(mm) > len(best[2])):
            best = (d, hi, mm)
    data = best[0] if best else next(iter(sheets.values()), None)
    if not data:
        return pd.DataFrame(columns=["ym", "목표", "실제"])
    hi = best[1] if best else 0
    header = [str(c).strip() for c in data[hi]]
    ym_cols = best[2] if best else {}
    result = {}
    if ym_cols:  # 가로형: 구분 x 월
        i_kind = next((j for j, c in enumerate(header) if "구분" in c or "목표" in c or "실제" in c), 0)
        for r in data[hi + 1:]:
            if not r or i_kind >= len(r) or r[i_kind] in (None, ""):
                continue
            kind = "목표" if "목표" in str(r[i_kind]) else ("실제" if ("실제" in str(r[i_kind]) or "실적" in str(r[i_kind])) else None)
            if kind is None:
                continue
            for j, ym in ym_cols.items():
                if j < len(r) and r[j] not in (None, ""):
                    result.setdefault(ym, {})[kind] = _f(r[j])
    else:  # 세로형: 월 | 목표 | 실제
        i_ym = 0
        i_t = next((j for j, c in enumerate(header) if "목표" in c), 1)
        i_a = next((j for j, c in enumerate(header) if "실제" in c or "실적" in c), 2)
        for r in data[1:]:
            ym = _norm_ym(r[i_ym]) if r and len(r) > i_ym else None
            if not ym:
                continue
            result.setdefault(ym, {})["목표"] = _f(r[i_t]) if len(r) > i_t else None
            result[ym]["실제"] = _f(r[i_a]) if len(r) > i_a else None
    rows = [{"ym": ym, "목표": v.get("목표"), "실제": v.get("실제")} for ym, v in result.items()]
    return pd.DataFrame(rows)


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def lfmall_value(lf_df, ym, kind="실제"):
    if lf_df is None or lf_df.empty:
        return None
    sub = lf_df[lf_df["ym"] == ym]
    if sub.empty:
        return None
    v = sub[kind].iloc[0]
    return None if pd.isna(v) else float(v)


# ======================================================================================
# 7) 업로드 양식(빈 템플릿) 생성 - 앱에서 다운로드 제공
# ======================================================================================
def build_target_template(year: int = 2026) -> bytes:
    """목표 업로드 양식 xlsx bytes. load_target()가 읽는 구조와 정확히 일치."""
    import io
    import openpyxl
    months = [f"{year}-{m:02d}" for m in range(1, 13)]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "목표"
    ws.append(["구분"] + months + ["TOTAL"])          # TOTAL 열은 읽을 때 무시됨
    ws.append([f"UV 목표 {year}"] + [0] * 12 + [None])   # UV: 지표 헤더행에 값 직접 입력
    for metric in ("인증자수", "당월인증거래액", "거래액"):
        ws.append([f"{metric} 목표 {year}"] + [None] * 12 + [None])  # 지표 헤더행
        for seg in ("신규", "win-back", "기존"):
            ws.append([f"  {seg}"] + [0] * 12 + [None])
    guide = wb.create_sheet("작성안내")
    for line in [
        "[목표 업로드 양식]",
        "· 월 열(2026-01 …)에 목표값을 입력하세요. TOTAL 열은 참고용(자동 무시).",
        "· 지표: UV / 인증자수 / 당월인증거래액 / 거래액",
        "· 각 지표 아래 신규 / win-back / 기존 행에 입력. UV는 헤더행에 바로 입력.",
        "· 거래액·당월인증거래액은 VAT 제외·원 단위.",
        "· 전체는 신규+윈백+기존 합으로 자동 계산됩니다.",
        "· 월 헤더는 2026-01 / 202601 / 2026.1 모두 인식합니다.",
    ]:
        guide.append([line])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_lfmall_template(year: int = 2026) -> bytes:
    """LF몰 전체거래액 업로드 양식 xlsx bytes. load_lfmall()가 읽는 구조와 일치."""
    import io
    import openpyxl
    months = [f"{year}-{m:02d}" for m in range(1, 13)]
    wb = openpyxl.Workbook()
    w = wb.active
    w.title = "LF몰전체거래액"
    w.append(["구분"] + months)
    w.append(["목표"] + [0] * 12)
    w.append(["실제"] + [0] * 12)
    guide = wb.create_sheet("작성안내")
    for line in [
        "[LF몰 전체거래액 업로드 양식]",
        "· 월별 LF몰 전체 거래액(VAT 제외·원)을 입력하세요.",
        "· '목표'/'실제' 2개 행. 비중 계산은 '실제' 기준입니다.",
    ]:
        guide.append([line])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
