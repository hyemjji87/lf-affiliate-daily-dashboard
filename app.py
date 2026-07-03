"""
LF몰 일반제휴 실적 대시보드 (Streamlit)

실행:  streamlit run app.py
필요 패키지: requirements.txt 참고

설계 요약 (대화에서 확정된 사양)
- 업로드 3종: ① 당해년도 실적  ② 전년도 인증 실적  ③ 전년도 거래 실적(분기, 복수 업로드 가능)
- 분석 기준일: (1) 특정 일자 복수선택  (2) 기준일 MTD(월초~기준일)
- 전년 동일자 매칭: 캘린더 동일 날짜 / 요일기준(52주 전) / 전년 특정 일자 복수선택 중 선택
- 모든 지표(UV·인증자수·거래액·고객수·객단가)의 모집단은 "선택 기간에 당월인증거래액이 발생한 제휴사"로 통일
- SSNGCD03(배너성 AF코드) UV 제외 옵션
- 총결제/순결제는 한 표에 섞지 않고 서브탭으로 분리하며, 표시 순서는 순결제 -> 총결제
- 전년비 표기: 상승 "▼", 하락 "△"(빨강) - 특정 회사 내부 컨벤션이라 이 순서를 그대로 사용
- 화면에는 전년비(%)만 표기하고, 전년도 원본 값은 엑셀 다운로드에만 포함
- 신규/윈백/기존 구분은 카테고리·브랜드 탭을 제외한 모든 탭에 적용
"""
import io
import os
import datetime as dt

import numpy as np
import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import lf_analysis as A

st.set_page_config(page_title="LF몰 일반제휴 실적", page_icon="📊", layout="wide")

# 일자별 목표(순결제 기준) 파일 - 연말까지 고정값이라 업로드 대신 리포에 고정 파일로 관리한다.
# 목표가 갱신되면 이 파일(target_data.xlsx)만 교체하면 된다 (별도 요청 시 반영).
TARGET_FILE = os.path.join(os.path.dirname(__file__), "target_data.xlsx")

# ----------------------------------------------------------------------------------
# 캐싱된 로더 (같은 파일이면 재계산하지 않음)
# ----------------------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def _load_single(file_bytes: bytes) -> dict:
    return A.classify_workbook(file_bytes)


@st.cache_data(show_spinner=False)
def _load_multi(list_of_bytes: tuple) -> dict:
    return A.merge_workbooks(list(list_of_bytes))


@st.cache_data(show_spinner=False)
def _load_target_from_disk(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        return pd.DataFrame()
    with open(path, "rb") as f:
        return A.load_target_file(f.read())


# ----------------------------------------------------------------------------------
# 포맷 유틸
# ----------------------------------------------------------------------------------
def fmt(n):
    if n is None or (isinstance(n, float) and pd.isna(n)):
        return "-"
    return f"{round(n):,}"


def sub(title: str):
    """st.subheader는 표 폰트(13px) 대비 지나치게 크므로, 표보다는 크되 절제된 크기의
    섹션 제목으로 대체."""
    st.markdown(f"<div style='font-size:16px;font-weight:700;margin:18px 0 6px;'>{title}</div>",
                unsafe_allow_html=True)


def pct_ratio(cur, prev):
    """엑셀/화면 공통으로 쓸 (표시문자열, 색상, 원시비율) 반환."""
    if prev is None or prev == 0 or pd.isna(prev) or cur is None or pd.isna(cur):
        if prev is None or (isinstance(prev, float) and pd.isna(prev)):
            return "신규", "accent", None
        return "-", "muted", None
    ratio = cur / prev - 1
    pct = round(ratio * 100)
    if pct < 0:
        return f"△{abs(pct)}%", "danger", ratio
    return f"▼{pct}%", "success", ratio


def target_ratio(cur, target):
    """목표비: 실적/목표 달성률(%) 숫자만 보여준다 - 전년비와 달리 ▼/△ 기호는 아예 쓰지 않는다
    (목표비는 절대 음수가 될 수 없어서 방향 기호가 의미가 없음). 100% 이상(달성·초과)이면 초록,
    100% 미만(미달)이면 빨강으로 색만 구분한다. 목표가 아예 없으면 '-'로 표기한다."""
    if (target is None or (isinstance(target, float) and pd.isna(target)) or target == 0
            or cur is None or (isinstance(cur, float) and pd.isna(cur))):
        return "-", "muted", None
    ratio = cur / target
    pct = round(ratio * 100)
    color = "success" if pct >= 100 else "danger"
    return f"{pct}%", color, ratio


COLOR_MAP = {"success": "#1a7f37", "danger": "#cf222e", "muted": "#6e7781", "accent": "#0969da"}


def style_delta_cols(df: pd.DataFrame, delta_cols: list | None = None):
    """Δ로 표기된 컬럼(문자열: '▼12%' / '△5%' / '신규' / '-')에 색을 입힌 Styler 반환.
    delta_cols=None 이면 컬럼 전체에 적용한다(멀티인덱스 컬럼 등 subset 지정이 번거로운 경우용 -
    _color는 ▼/△/신규 표기가 아닌 값에는 항상 빈 스타일을 반환하므로 전체 적용해도 안전하다)."""
    def _color(v):
        if isinstance(v, str) and v.startswith("▼"):
            return f"color:{COLOR_MAP['success']};font-weight:600"
        if isinstance(v, str) and v.startswith("△"):
            return f"color:{COLOR_MAP['danger']};font-weight:600"
        if v == "신규":
            return f"color:{COLOR_MAP['accent']};font-weight:600"
        return ""
    if delta_cols is None:
        return df.style.map(_color)
    return df.style.map(_color, subset=delta_cols)


def style_with_color_grid(df: pd.DataFrame, color_grid: dict):
    """목표비처럼 표시 기호(▼)만으로는 색상(달성=초록/미달=빨강)을 구분할 수 없는 컬럼용.
    color_grid = {컬럼명: [이 컬럼의 행별 색상키('success'/'danger'/'accent'/None), ...]}.
    지정 안 한 컬럼은 색을 입히지 않는다."""
    def _apply(_df):
        styles = pd.DataFrame("", index=_df.index, columns=_df.columns)
        for col, keys in color_grid.items():
            if col not in styles.columns:
                continue
            for i, key in enumerate(keys):
                if key in ("success", "danger", "accent"):
                    styles.iloc[i, styles.columns.get_loc(col)] = f"color:{COLOR_MAP[key]};font-weight:600"
        return styles
    return df.style.apply(_apply, axis=None)


def render_grouped_table(header_groups: list, rows: list, color_rows: list | None = None, first_col_label: str = "일자"):
    """일자별 탭처럼 '상위 카테고리(전체/신규/윈백/기존) + 하위 지표(실적/전년비/목표비)'를
    2단 헤더로 묶어서 보여주는 HTML 표. 첫 컬럼(일자)은 1·2행을 세로 병합하고, 모든 셀은
    가운데 정렬한다.
    - header_groups: [(그룹명, [하위컬럼명, ...]), ...]
    - rows: [[일자표시값, 값1, 값2, ...], ...] (값은 이미 fmt()된 문자열)
    - color_rows: rows와 동일한 shape(첫 컬럼 제외)으로 각 셀의 색상키('success'/'danger'/'accent'/None)를
      담은 리스트. None이면 색상 없이 렌더링.
    """
    ths_top = ['<th rowspan="2" style="text-align:center;vertical-align:middle;'
               'border:1px solid #e1e4e8;padding:6px 10px;background:#f6f8fa;">'
               f'{first_col_label}</th>']
    for g, subs in header_groups:
        ths_top.append(f'<th colspan="{len(subs)}" style="text-align:center;'
                        f'border:1px solid #e1e4e8;padding:6px 10px;background:#f6f8fa;">{g}</th>')
    ths_sub = []
    for g, subs in header_groups:
        for s in subs:
            ths_sub.append(f'<th style="text-align:center;border:1px solid #e1e4e8;'
                            f'padding:4px 10px;background:#fafbfc;font-weight:500;">{s}</th>')

    body_rows = []
    for ri, row in enumerate(rows):
        cells = [f'<td style="text-align:center;border:1px solid #e1e4e8;'
                 f'padding:4px 10px;white-space:nowrap;">{row[0]}</td>']
        for ci, v in enumerate(row[1:]):
            style = "text-align:center;border:1px solid #e1e4e8;padding:4px 10px;white-space:nowrap;"
            if color_rows is not None:
                key = color_rows[ri][ci]
                if key in ("success", "danger", "accent"):
                    style += f"color:{COLOR_MAP[key]};font-weight:600;"
            cells.append(f'<td style="{style}">{v}</td>')
        body_rows.append(f'<tr>{"".join(cells)}</tr>')

    html = (
        '<div style="overflow-x:auto;">'
        '<table style="border-collapse:collapse;width:100%;font-size:13px;">'
        f'<thead><tr>{"".join(ths_top)}</tr><tr>{"".join(ths_sub)}</tr></thead>'
        f'<tbody>{"".join(body_rows)}</tbody>'
        '</table></div>'
    )
    return html


EXCEL_PCT_FMT = '"▼"0%;[Red]"△"0%;0%'


def to_excel_bytes(sheets: dict, pct_cols_by_sheet: dict | None = None) -> bytes:
    """{시트명: DataFrame} -> xlsx 바이트. pct_cols_by_sheet={'시트명':[컬럼명,...]} 이면
    해당 컬럼에 회사 컨벤션 숫자서식('▼0%;[빨강]△0%')을 적용한다(원시 비율 값이 들어있어야 함)."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, df in sheets.items():
            safe_name = name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
            ws = writer.sheets[safe_name]
            if pct_cols_by_sheet and name in pct_cols_by_sheet:
                for col in pct_cols_by_sheet[name]:
                    if col in df.columns:
                        col_idx = df.columns.get_loc(col) + 1  # 1-based
                        col_letter = ws.cell(row=1, column=col_idx).column_letter
                        for r in range(2, len(df) + 2):
                            ws[f"{col_letter}{r}"].number_format = EXCEL_PCT_FMT
    return buf.getvalue()


_WEEKDAY_KR = ["월", "화", "수", "목", "금", "토", "일"]
_THIN = Side(style="thin", color="D0D7DE")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TITLE_FONT = Font(bold=True, size=12)
_HEADER_FONT = Font(bold=True, size=10)
_HEADER_FILL = PatternFill("solid", fgColor="F0F3F6")
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _pct_font(is_good: bool) -> Font:
    return Font(color=("1A7F37" if is_good else "CF222E"), bold=True)


def build_exec_report_excel(
    sel_dates, ly_all_dates, yoy_mode_label,
    cur_all, prev_all, cur_cert, prev_cert,
    cert_cur, cert_prev,
    all_cur_d, all_ly_d, cert_cur_d, cert_ly_d,
    pf_cur, pf_ly, target_df,
) -> bytes:
    """임원보고용 종합 리포트: ① 지표별 실적 요약 ② 일자별/신규·윈백·기존별 거래액 실적
    (전년비·목표달성률·MIX 포함) ③ 제휴사별 인증/거래액 실적 (전년비 포함), 순결제 기준 고정.
    사용자가 제공한 참조 엑셀 양식을 따라 구성했다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "종합 리포트"
    days_sorted = sorted(sel_dates)
    n_days = len(days_sorted) or 1
    r = 1

    def title_row(text):
        nonlocal r
        ws.cell(r, 1, text).font = _TITLE_FONT
        r += 2

    def style_header(row):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row, c)
            if cell.value is not None or c <= 40:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _CENTER
                cell.border = _BORDER

    def style_row(row, ncols, first_left=True):
        for c in range(1, ncols + 1):
            cell = ws.cell(row, c)
            cell.border = _BORDER
            cell.alignment = _LEFT if (c == 1 and first_left) else _CENTER

    # ---------------- 1) 지표별 실적 요약 ----------------
    title_row("■ 지표별 실적 요약")
    hdr1 = r
    ws.cell(hdr1, 1, "구분")
    ws.merge_cells(start_row=hdr1, start_column=1, end_row=hdr1 + 1, end_column=1)
    groups1 = [("거래액", 2), ("당월인증 거래액", 2), ("인증회원수", 2)]
    col = 2
    for name, span in groups1:
        ws.cell(hdr1, col, name)
        ws.merge_cells(start_row=hdr1, start_column=col, end_row=hdr1, end_column=col + span - 1)
        col += span
    for i, lbl in enumerate(["MTD", "일평균"] * 3):
        ws.cell(hdr1 + 1, 2 + i, lbl)
    total_cols1 = 1 + sum(s for _, s in groups1)
    style_header(hdr1)
    style_header(hdr1 + 1)
    r = hdr1 + 2
    all_net_total = cur_all["전체"]["net"]
    cert_net_total = cur_cert["전체"]["net"]
    cert_cnt_total = cert_cur["전체"]
    vals1 = [all_net_total, all_net_total / n_days, cert_net_total, cert_net_total / n_days,
             cert_cnt_total, cert_cnt_total / n_days]
    ws.cell(r, 1, "")
    for i, v in enumerate(vals1):
        cell = ws.cell(r, 2 + i, round(v))
        cell.number_format = "#,##0"
    style_row(r, total_cols1)
    r += 2

    # ---------------- 2) 일자별/신규·윈백·기존별 거래액 실적 ----------------
    title_row("■ 일자별/신규·윈백·기존별 거래액 실적 (순결제 기준)")
    hdr2 = r
    ws.cell(hdr2, 1, "구분")
    ws.merge_cells(start_row=hdr2, start_column=1, end_row=hdr2 + 1, end_column=1)
    for i, d in enumerate(days_sorted):
        c = 2 + i
        ws.cell(hdr2, c, d.strftime("%Y-%m-%d"))
        ws.cell(hdr2 + 1, c, _WEEKDAY_KR[d.weekday()])
    base = 2 + len(days_sorted)
    extra_labels = ["MTD 실적", "전년비", "목표달성률", "MIX"]
    for i, lbl in enumerate(extra_labels):
        ws.cell(hdr2, base + i, lbl)
        ws.merge_cells(start_row=hdr2, start_column=base + i, end_row=hdr2 + 1, end_column=base + i)
    total_cols2 = base + len(extra_labels) - 1
    style_header(hdr2)
    style_header(hdr2 + 1)
    r = hdr2 + 2

    def write_metric_rows(label_total, daily_df, ly_daily_df, target_group):
        nonlocal r
        status_defs = [("전체", "total", label_total), ("신규", "신규", "신규"),
                        ("WIN-BACK", "윈백", "WIN-BACK"), ("기존", "기존", "기존")]
        total_mtd = None
        for key, tgt_status, label in status_defs:
            row_label = label if key == "전체" else f"    {label}"
            daily_vals = []
            for d in days_sorted:
                match = daily_df[daily_df["일자"] == d]
                v = float(match.iloc[0][f"{key}_net"]) if not match.empty else 0.0
                daily_vals.append(v)
            mtd = sum(daily_vals)
            if key == "전체":
                total_mtd = mtd
            ly_sum = float(ly_daily_df[f"{key}_net"].sum()) if (ly_daily_df is not None and not ly_daily_df.empty) else None
            _, d_color, d_ratio = pct_ratio(mtd, ly_sum)
            tgt_v = A.target_sum(target_df, sel_dates, target_group, tgt_status) if not target_df.empty else None
            _, t_color, t_ratio = target_ratio(mtd, tgt_v)
            mix = (mtd / total_mtd) if total_mtd else 0

            ws.cell(r, 1, row_label)
            for i, v in enumerate(daily_vals):
                cell = ws.cell(r, 2 + i, round(v))
                cell.number_format = "#,##0"
            mtd_cell = ws.cell(r, base, round(mtd))
            mtd_cell.number_format = "#,##0"
            ly_cell = ws.cell(r, base + 1, round(d_ratio, 4) if d_ratio is not None else None)
            ly_cell.number_format = EXCEL_PCT_FMT
            if d_color in ("success", "danger"):
                ly_cell.font = _pct_font(d_color == "success")
            tgt_cell = ws.cell(r, base + 2, round(t_ratio, 4) if t_ratio is not None else None)
            tgt_cell.number_format = "0%"
            if t_color in ("success", "danger"):
                tgt_cell.font = _pct_font(t_color == "success")
            mix_cell = ws.cell(r, base + 3, round(mix, 4))
            mix_cell.number_format = "0%"
            style_row(r, total_cols2)
            r += 1

    write_metric_rows("거래액", all_cur_d, all_ly_d, "all")
    write_metric_rows("당월 인증거래액", cert_cur_d, cert_ly_d, "cert")
    r += 1

    # ---------------- 3) 제휴사별 인증/거래액 실적 ----------------
    title_row("■ 제휴사별 인증/거래액 실적 (순결제 기준)")
    hdr3 = r
    ws.cell(hdr3, 1, "구분")
    ws.merge_cells(start_row=hdr3, start_column=1, end_row=hdr3 + 1, end_column=1)
    groups3 = [("거래액", 2), ("당월인증 거래액", 2), ("인증회원수", 2)]
    col = 2
    for name, span in groups3:
        ws.cell(hdr3, col, name)
        ws.merge_cells(start_row=hdr3, start_column=col, end_row=hdr3, end_column=col + span - 1)
        col += span
    for i, lbl in enumerate(["MTD", "전년비"] * 3):
        ws.cell(hdr3 + 1, 2 + i, lbl)
    total_cols3 = 1 + sum(s for _, s in groups3)
    style_header(hdr3)
    style_header(hdr3 + 1)
    r = hdr3 + 2

    def write_pair_row(label, triples, indent=False):
        nonlocal r
        ws.cell(r, 1, ("    " if indent else "") + label)
        c = 2
        for mtd_v, prev_v in triples:
            mtd_cell = ws.cell(r, c, round(mtd_v))
            mtd_cell.number_format = "#,##0"
            _, d_color, d_ratio = pct_ratio(mtd_v, prev_v)
            pc = ws.cell(r, c + 1, round(d_ratio, 4) if d_ratio is not None else None)
            pc.number_format = EXCEL_PCT_FMT
            if d_color in ("success", "danger"):
                pc.font = _pct_font(d_color == "success")
            c += 2
        style_row(r, total_cols3)
        r += 1

    write_pair_row("전체", [
        (all_net_total, prev_all["전체"]["net"] if prev_all["전체"]["net"] is not None else None),
        (cert_net_total, prev_cert["전체"]["net"] if prev_cert["전체"]["net"] is not None else None),
        (cert_cnt_total, cert_prev.get("전체")),
    ])
    pf_ly_idx = pf_ly.set_index("제휴사") if pf_ly is not None else None
    for _, prow in pf_cur.iterrows():
        p = pf_ly_idx.loc[prow["제휴사"]] if (pf_ly_idx is not None and prow["제휴사"] in pf_ly_idx.index) else None
        write_pair_row(prow["제휴사"], [
            (prow["all_net"], p["all_net"] if p is not None else None),
            (prow["cert_net"], p["cert_net"] if p is not None else None),
            (prow["인증수"], p["인증수"] if p is not None else None),
        ], indent=True)

    ws.column_dimensions["A"].width = 22
    for c in range(2, max(total_cols2, total_cols3) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.freeze_panes = "B1"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ----------------------------------------------------------------------------------
# 사이드바: 파일 업로드 & 분석 설정
# ----------------------------------------------------------------------------------
st.sidebar.header("📁 Step 1. 파일 업로드")
f_cur = st.sidebar.file_uploader("① 당해년도 실적", type=["xlsx"], key="f_cur")
f_ly_cert = st.sidebar.file_uploader("② 전년도 인증 실적", type=["xlsx"], key="f_ly_cert")
f_ly_amt = st.sidebar.file_uploader(
    "③ 전년도 거래 실적 (분기별, 여러 파일 업로드 가능)",
    type=["xlsx"], accept_multiple_files=True, key="f_ly_amt",
)

st.sidebar.divider()
st.sidebar.header("⚙ Step 2. 분석 설정")

if f_cur is None:
    st.sidebar.info("① 당해년도 실적 파일을 올리면 분석 설정이 열립니다.")
    st.title("📊 LF몰 일반제휴 실적")
    st.info("왼쪽 사이드바에서 ① 당해년도 실적 파일을 먼저 업로드해 주세요.")
    st.stop()

cur_data = _load_single(f_cur.getvalue())
if cur_data["amt"].empty:
    st.error("업로드한 파일에서 '인증거래액' 형태의 시트를 찾지 못했습니다. 파일을 확인해 주세요.")
    st.stop()

available_dates = sorted(cur_data["amt"]["정산일시일"].dropna().dt.date.unique().tolist())
min_d, max_d = available_dates[0], available_dates[-1]

mode = st.sidebar.radio("분석 기준일", ["특정 일자 복수선택", "기준일 MTD"], index=1)

if mode == "특정 일자 복수선택":
    sel_dates = st.sidebar.multiselect(
        "날짜 선택 (드롭다운, 여러 개 선택 가능)",
        options=available_dates, default=available_dates,
        format_func=lambda d: d.strftime("%Y-%m-%d"),
    )
else:
    base_date = st.sidebar.date_input(
        "MTD 마감일", value=max_d, min_value=min_d, max_value=max_d,
    )
    month_start = base_date.replace(day=1)
    sel_dates = [d for d in available_dates if month_start <= d <= base_date]

if not sel_dates:
    st.warning("분석할 날짜를 하나 이상 선택해 주세요.")
    st.stop()

excl_ssng = st.sidebar.checkbox("SSNGCD03(삼성배너) UV 제외", value=False)
yoy_mode_label = st.sidebar.radio(
    "전년 동일자 매칭",
    ["캘린더 동일 날짜", "요일기준(52주 전)", "전년 특정 일자 복수선택"],
    index=0,
)
ly_manual_dates = None
if yoy_mode_label == "전년 특정 일자 복수선택":
    yoy_mode = "manual"
    ly_candidates = A.ly_month_candidates(sel_dates)
    _default_ly = [d for d in A.ly_dates(sel_dates, "calendar") if d in ly_candidates]
    ly_manual_dates = st.sidebar.multiselect(
        "전년도 비교 일자 선택 (해당 월 전체 중 복수선택 가능)",
        options=ly_candidates, default=_default_ly,
        format_func=lambda d: d.strftime("%Y-%m-%d"),
    )
    if not ly_manual_dates:
        st.sidebar.caption("⚠ 전년도 비교 일자를 선택하지 않으면 전년비가 계산되지 않습니다.")
else:
    yoy_mode = "calendar" if yoy_mode_label == "캘린더 동일 날짜" else "weekday"
exclude_af = "SSNGCD03" if excl_ssng else None

# ----------------------------------------------------------------------------------
# 전년도 데이터 로딩 (선택사항 - 없으면 전년비는 "-"로 표기)
# ----------------------------------------------------------------------------------
ly_cert_data = _load_single(f_ly_cert.getvalue()) if f_ly_cert else {"amt": pd.DataFrame(), "uv": pd.DataFrame(), "cert": pd.DataFrame()}
ly_amt_bytes = tuple(f.getvalue() for f in f_ly_amt) if f_ly_amt else tuple()
ly_amt_data = _load_multi(ly_amt_bytes) if ly_amt_bytes else {"amt": pd.DataFrame(), "uv": pd.DataFrame(), "cert": pd.DataFrame()}
target_df = _load_target_from_disk(TARGET_FILE)

ly_all_dates = ly_manual_dates if yoy_mode == "manual" else A.ly_dates(sel_dates, yoy_mode)

# [주의] qual_partners(당월인증거래액>0인 제휴사)는 "제휴사별 실적" 탭에 어떤 제휴사를
# 리스팅할지 정하는 기준일 뿐이다. UV/인증자수/거래액 등 다른 모든 탭의 집계는 이 리스트로
# 제한하지 않고 업로드된 전체 제휴사(partners=None) 기준으로 계산한다.
qual_partners = A.qualifying_partners(cur_data["amt"], sel_dates)

# ----------------------------------------------------------------------------------
# 헤더
# ----------------------------------------------------------------------------------
warn_bits = []
if f_ly_cert is None:
    warn_bits.append("전년도 인증 실적 미업로드(UV·인증자수 전년비 불가)")
if not f_ly_amt:
    warn_bits.append("전년도 거래 실적 미업로드(거래액 전년비 불가)")

st.title("📊 LF몰 일반제휴 실적")
period_txt = f"{sel_dates[0]:%Y-%m-%d} ~ {sel_dates[-1]:%Y-%m-%d}" if len(sel_dates) > 1 else f"{sel_dates[0]:%Y-%m-%d}"
st.caption(
    f"분석기간 {period_txt} · UV·인증자수·거래액 등은 전체 제휴사 기준 · "
    f"제휴사별 실적 탭 리스팅 대상 {len(qual_partners)}개(당월인증거래액 발생 기준) · "
    f"전년 매칭: {yoy_mode_label}" + ("" if not warn_bits else " · ⚠ " + " / ".join(warn_bits))
)

json_col, _ = st.columns([1, 5])
with json_col:
    pass  # JSON 내보내기 버튼은 각 탭 하단에서 개별 제공 (탭별 데이터가 다르므로)

tab_ov, tab_daily, tab_partner, tab_cat, tab_brand = st.tabs(
    ["개요", "일자별", "제휴사별 실적", "카테고리별 실적", "브랜드별 실적"]
)

# ====================================================================================
# 개요 탭
# ====================================================================================
with tab_ov:
    uv_cur = A.uv_total(cur_data["uv"], sel_dates, None, exclude_af=exclude_af)
    uv_prev = A.uv_total(ly_cert_data["uv"], ly_all_dates, None) if not ly_cert_data["uv"].empty else None
    cert_cur = A.cert_summary(cur_data["cert"], sel_dates, None)
    cert_prev = A.cert_summary(ly_cert_data["cert"], ly_all_dates, None) if not ly_cert_data["cert"].empty else {"전체": None, "기존": None, "WIN-BACK": None, "신규": None}

    c1, c2 = st.columns(2)
    with c1:
        st.metric("UV", fmt(uv_cur), pct_ratio(uv_cur, uv_prev)[0])
    with c2:
        st.metric("인증자수", fmt(cert_cur["전체"]), pct_ratio(cert_cur["전체"], cert_prev["전체"])[0])

    sub("인증자수 구성 (비중은 기준일 기준)")
    rows = []
    for k, label in [("신규", "신규"), ("WIN-BACK", "윈백"), ("기존", "기존")]:
        share = f"{cert_cur[k] / cert_cur['전체'] * 100:.1f}%" if cert_cur["전체"] else "-"
        d, _, _ = pct_ratio(cert_cur[k], cert_prev.get(k))
        rows.append([label, fmt(cert_cur[k]), share, d])
    df_certcomp = pd.DataFrame(rows, columns=["구분", "기준일", "비중", "전년비"])
    st.dataframe(style_delta_cols(df_certcomp, ["전년비"]), hide_index=True, use_container_width=True)

    paytype = st.radio("결제 구분", ["순결제", "총결제"], horizontal=True, key="ov_paytype")
    pt = "net" if paytype == "순결제" else "tot"

    def _amt_block(label, title):
        cur = A.amt_summary(cur_data["amt"], sel_dates, None, cert_only=(label == "당월인증"))
        prev = (A.amt_summary(ly_amt_data["amt"], ly_all_dates, None, cert_only=(label == "당월인증"))
                if not ly_amt_data["amt"].empty else {s: {"tot": None, "net": None, "cust_tot": None, "cust_net": None} for s in ["전체"] + A.STATUSES})
        cur_total = cur["전체"][pt]
        prev_total = prev["전체"][pt]
        d_all, d_all_color, _ = pct_ratio(cur_total, prev_total)
        show_target = pt == "net" and not target_df.empty
        group = "cert" if label == "당월인증" else "all"
        d_all_css = f"color:{COLOR_MAP.get(d_all_color, '#000')};font-weight:600" if d_all_color in COLOR_MAP else ""
        header = (f"**{title}**  \n### {fmt(cur_total)}  "
                  f"<span style='font-size:13px;{d_all_css}'>전년비 {d_all}</span>")
        if show_target:
            tgt_total = A.target_sum(target_df, sel_dates, group, "total")
            t_all, t_all_color, _ = target_ratio(cur_total, tgt_total)
            t_all_css = f"color:{COLOR_MAP.get(t_all_color, '#000')};font-weight:600" if t_all_color in COLOR_MAP else ""
            header += f"  <span style='font-size:13px;{t_all_css}'>목표비 {t_all}</span>"
        st.markdown(header, unsafe_allow_html=True)
        rows = []
        delta_colors, target_colors = [], []
        status_target_key = {"신규": "신규", "WIN-BACK": "윈백", "기존": "기존"}
        for k, lbl in [("신규", "신규"), ("WIN-BACK", "윈백"), ("기존", "기존")]:
            c = cur[k][pt]
            p = prev[k][pt]
            share = f"{c / cur_total * 100:.1f}%" if cur_total else "-"
            d, d_color, _ = pct_ratio(c, p)
            row = [lbl, fmt(c), share, d]
            delta_colors.append(d_color)
            if show_target:
                tgt_v = A.target_sum(target_df, sel_dates, group, status_target_key[k])
                t, t_color, _ = target_ratio(c, tgt_v)
                row.append(t)
                target_colors.append(t_color)
            rows.append(row)
        cols = ["구분", "기준일", "비중", "전년비"] + (["목표비"] if show_target else [])
        df = pd.DataFrame(rows, columns=cols)
        color_grid = {"전년비": delta_colors}
        if show_target:
            color_grid["목표비"] = target_colors
        st.dataframe(style_with_color_grid(df, color_grid), hide_index=True, use_container_width=True)
        return cur, prev

    cur_all, prev_all = _amt_block("전체", "전체거래액")
    cur_cert, prev_cert = _amt_block("당월인증", "당월인증거래액")

    def _cust_aov_table(cur, prev, title):
        sub(title)
        cust_c, cust_p = cur["전체"][f"cust_{pt}"], prev["전체"][f"cust_{pt}"]
        aov_c = cur["전체"][pt] / cust_c if cust_c else 0
        aov_p = (prev["전체"][pt] / cust_p) if cust_p else None
        df = pd.DataFrame([
            ["고객수", fmt(cust_c), pct_ratio(cust_c, cust_p)[0]],
            ["객단가", fmt(aov_c), pct_ratio(aov_c, aov_p)[0]],
        ], columns=["구분", "기준일", "전년비"])
        st.dataframe(style_delta_cols(df, ["전년비"]), hide_index=True, use_container_width=True)

    _cust_aov_table(cur_all, prev_all, "전체 고객수 · 객단가")
    _cust_aov_table(cur_cert, prev_cert, "당월인증 고객수 · 객단가")

# ====================================================================================
# 일자별 탭 (화면 = 값 + 전년비만, 전년 원본값은 엑셀에만)
# ====================================================================================
with tab_daily:
    daily_uc = A.daily_uv_cert_table(cur_data["uv"], cur_data["cert"], sel_dates, None, exclude_af=exclude_af)
    daily_uc_ly = (A.daily_uv_cert_table(ly_cert_data["uv"], ly_cert_data["cert"], ly_all_dates, None)
                   if (not ly_cert_data["uv"].empty or not ly_cert_data["cert"].empty) else None)

    sub("UV")
    uv_rows = []
    for i, d in enumerate(sorted(sel_dates)):
        prev_v = daily_uc_ly.iloc[i]["UV"] if daily_uc_ly is not None and i < len(daily_uc_ly) else None
        uv_rows.append([d, fmt(daily_uc.iloc[i]["UV"]), pct_ratio(daily_uc.iloc[i]["UV"], prev_v)[0]])
    df_uv = pd.DataFrame(uv_rows, columns=["일자", "UV", "전년비"])
    st.dataframe(style_delta_cols(df_uv, ["전년비"]), hide_index=True, use_container_width=True)

    sub("인증자수 (전체·신규·윈백·기존)")
    cert_rows, cert_color_rows = [], []
    for i, d in enumerate(sorted(sel_dates)):
        r = daily_uc.iloc[i]
        pr = daily_uc_ly.iloc[i] if daily_uc_ly is not None and i < len(daily_uc_ly) else None
        row = [d.strftime("%Y-%m-%d")]
        color_row = []
        for col in ["인증_전체", "인증_신규", "인증_윈백", "인증_기존"]:
            pv = pr[col] if pr is not None else None
            d_disp, d_color, _ = pct_ratio(r[col], pv)
            row += [fmt(r[col]), d_disp]
            color_row += [None, d_color]
        cert_rows.append(row)
        cert_color_rows.append(color_row)
    cert_groups = [(s, ["인증수", "전년비"]) for s in ["전체", "신규", "윈백", "기존"]]
    st.markdown(render_grouped_table(cert_groups, cert_rows, cert_color_rows), unsafe_allow_html=True)

    paytype_d = st.radio("결제 구분", ["순결제", "총결제"], horizontal=True, key="daily_paytype")
    ptd = "net" if paytype_d == "순결제" else "tot"

    def _daily_amt_block(cert_only, title):
        cur_d = A.daily_amt_table(cur_data["amt"], sel_dates, None, cert_only=cert_only)
        ly_d = (A.daily_amt_table(ly_amt_data["amt"], ly_all_dates, None, cert_only=cert_only)
                if not ly_amt_data["amt"].empty else None)
        sub(f"{title} (전체·신규·윈백·기존)")
        show_target = ptd == "net" and not target_df.empty
        group = "cert" if cert_only else "all"
        status_target_key = {"전체": "total", "신규": "신규", "WIN-BACK": "윈백", "기존": "기존"}
        rows, color_rows = [], []
        for i in range(len(cur_d)):
            r = cur_d.iloc[i]
            pr = ly_d.iloc[i] if ly_d is not None and i < len(ly_d) else None
            row = [r["일자"].strftime("%Y-%m-%d")]
            color_row = []
            for s in ["전체"] + A.STATUSES:
                c = r[f"{s}_{ptd}"]
                p = pr[f"{s}_{ptd}"] if pr is not None else None
                d_disp, d_color, _ = pct_ratio(c, p)
                row += [fmt(c), d_disp]
                color_row += [None, d_color]
                if show_target:
                    tgt_v = A.target_lookup(target_df, r["일자"], group, status_target_key[s])
                    t_disp, t_color, _ = target_ratio(c, tgt_v)
                    row.append(t_disp)
                    color_row.append(t_color)
            rows.append(row)
            color_rows.append(color_row)
        subcols = ["실적", "전년비"] + (["목표비"] if show_target else [])
        groups = [(s, subcols) for s in ["전체", "신규", "윈백", "기존"]]
        st.markdown(render_grouped_table(groups, rows, color_rows), unsafe_allow_html=True)
        return cur_d, ly_d

    all_cur_d, all_ly_d = _daily_amt_block(False, "전체거래액")
    cert_cur_d, cert_ly_d = _daily_amt_block(True, "당월인증거래액")

    def _daily_cust_aov(cur_d, ly_d, title):
        sub(title)
        rows = []
        for i in range(len(cur_d)):
            r = cur_d.iloc[i]
            pr = ly_d.iloc[i] if ly_d is not None and i < len(ly_d) else None
            cust_c = r[f"전체_cust_{ptd}"]
            aov_c = r[f"전체_{ptd}"] / cust_c if cust_c else 0
            cust_p = pr[f"전체_cust_{ptd}"] if pr is not None else None
            aov_p = (pr[f"전체_{ptd}"] / cust_p) if (pr is not None and cust_p) else None
            rows.append([r["일자"], fmt(cust_c), pct_ratio(cust_c, cust_p)[0], fmt(aov_c), pct_ratio(aov_c, aov_p)[0]])
        df = pd.DataFrame(rows, columns=["일자", "고객수", "고객수 전년비", "객단가", "객단가 전년비"])
        st.dataframe(style_delta_cols(df, ["고객수 전년비", "객단가 전년비"]), hide_index=True, use_container_width=True)

    _daily_cust_aov(cert_cur_d, cert_ly_d, "고객수 · 객단가 (당월인증 기준)")
    _daily_cust_aov(all_cur_d, all_ly_d, "고객수 · 객단가 (전체거래액 기준)")
    st.caption("전년도 원본 값은 화면에 표시하지 않으며, 아래 엑셀 다운로드에는 함께 포함됩니다.")

    excel_bytes = to_excel_bytes({
        "UV_인증자수": daily_uc.assign(UV_전년=daily_uc_ly["UV"] if daily_uc_ly is not None else np.nan),
        "전체거래액": all_cur_d,
        "당월인증거래액": cert_cur_d,
    })
    st.download_button("⤓ 엑셀 다운로드 (일자별, 전년도 원본 포함)", excel_bytes, file_name="일자별_실적.xlsx")

# ====================================================================================
# 제휴사별 실적 탭
# ====================================================================================
with tab_partner:
    paytype_p = st.radio("결제 구분", ["순결제", "총결제"], horizontal=True, key="partner_paytype")
    ptp = "net" if paytype_p == "순결제" else "tot"

    pf_cur = A.partner_full_table(cur_data["amt"], cur_data["uv"], cur_data["cert"], sel_dates, qual_partners, exclude_af=exclude_af)
    pf_ly = (A.partner_full_table(ly_amt_data["amt"], ly_cert_data["uv"], ly_cert_data["cert"], ly_all_dates, None)
             if not ly_amt_data["amt"].empty else None)
    ly_map = {r["제휴사"]: r for _, r in pf_ly.iterrows()} if pf_ly is not None else {}

    sub(f"전체 제휴사 (당월인증거래액 발생 기준, 내림차순 · {len(pf_cur)}개)")
    rows = []
    for _, r in pf_cur.iterrows():
        p = ly_map.get(r["제휴사"])
        cert_c, all_c = r[f"cert_{ptp}"], r[f"all_{ptp}"]
        cust_c = r[f"cust_{ptp}"]
        aov_c = all_c / cust_c if cust_c else 0
        if p is not None:
            cert_p, all_p, cust_p = p[f"cert_{ptp}"], p[f"all_{ptp}"], p[f"cust_{ptp}"]
            aov_p = all_p / cust_p if cust_p else None
            uv_p, cert_cnt_p = p["UV"], p["인증수"]
        else:
            cert_p = all_p = cust_p = aov_p = uv_p = cert_cnt_p = None
        rows.append([
            r["제휴사"], fmt(r["UV"]), pct_ratio(r["UV"], uv_p)[0],
            fmt(r["인증수"]), pct_ratio(r["인증수"], cert_cnt_p)[0],
            fmt(cert_c), pct_ratio(cert_c, cert_p)[0],
            fmt(all_c), pct_ratio(all_c, all_p)[0],
            fmt(cust_c), pct_ratio(cust_c, cust_p)[0],
            fmt(aov_c), pct_ratio(aov_c, aov_p)[0],
        ])
    cols = ["제휴사", "UV", "UV 전년비", "인증수", "인증수 전년비", "당월인증거래액", "당월인증거래액 전년비",
            "전체거래액", "전체거래액 전년비", "고객수", "고객수 전년비", "객단가", "객단가 전년비"]
    partner_delta_cols = ["UV 전년비", "인증수 전년비", "당월인증거래액 전년비", "전체거래액 전년비", "고객수 전년비", "객단가 전년비"]
    df_partner = pd.DataFrame(rows, columns=cols)
    st.dataframe(style_delta_cols(df_partner, partner_delta_cols), hide_index=True, use_container_width=True)
    st.caption("전년비는 전년도 동일 제휴사 코호트 기준(전년도 데이터가 없으면 \"신규\" 표기).")

    pf_ly_idx = pf_ly.set_index("제휴사") if pf_ly is not None else None

    def _with_ly(df, cols):
        out = {}
        for c in cols:
            if pf_ly_idx is not None and c in pf_ly_idx.columns:
                out[f"전년_{c}"] = df["제휴사"].map(pf_ly_idx[c])
            else:
                out[f"전년_{c}"] = np.nan
        return df.assign(**out)

    export_sheets = {
        "제휴사별_순결제": _with_ly(pf_cur, ["cert_net", "all_net", "cust_net"]),
        "제휴사별_총결제": _with_ly(pf_cur, ["cert_tot", "all_tot", "cust_tot"]),
    }

    if not qual_partners:
        st.info("선택한 기간에 당월인증거래액이 발생한 제휴사가 없어 리스팅할 제휴사가 없습니다.")
    else:
        sub("제휴사 선택 → 일자별 상세")
        sel_partner = st.selectbox("제휴사", qual_partners)
        pd_cur = A.partner_daily_table(cur_data["amt"], cur_data["uv"], cur_data["cert"], sel_dates, sel_partner, exclude_af=exclude_af)
        show_cols = ["일자", "UV", "인증_전체",
                     f"당월인증_{ptp}", f"전체거래액_{ptp}", f"고객수_{ptp}"]
        disp = pd_cur[show_cols].rename(columns={
            "인증_전체": "인증수", f"당월인증_{ptp}": "당월인증거래액",
            f"전체거래액_{ptp}": "전체거래액", f"고객수_{ptp}": "고객수"})
        disp["객단가"] = (disp["전체거래액"] / disp["고객수"].replace(0, np.nan)).fillna(0)
        st.dataframe(disp, hide_index=True, use_container_width=True)
        export_sheets[f"{sel_partner}_일자별"] = pd_cur

    excel_bytes = to_excel_bytes(export_sheets)
    st.download_button("⤓ 엑셀 다운로드 (제휴사별, 전년도 원본 포함)", excel_bytes, file_name="제휴사별_실적.xlsx")

# ====================================================================================
# 카테고리별 실적 탭
# ====================================================================================
with tab_cat:
    paytype_c = st.radio("결제 구분", ["순결제", "총결제"], horizontal=True, key="cat_paytype")
    ptc = "net" if paytype_c == "순결제" else "tot"

    cat_cert = A.group_amt_table(cur_data["amt"], sel_dates, None, "물리대카테", cert_only=True).rename(columns={"tot": "cert_tot", "net": "cert_net"})
    cat_all = A.group_amt_table(cur_data["amt"], sel_dates, None, "물리대카테", cert_only=False).rename(columns={"tot": "all_tot", "net": "all_net"})
    cat = pd.merge(cat_cert, cat_all, on="물리대카테", how="outer").fillna(0)
    cat = cat.sort_values("cert_tot", ascending=False).reset_index(drop=True)

    cat_ly = None
    if not ly_amt_data["amt"].empty:
        cat_cert_ly = A.group_amt_table(ly_amt_data["amt"], ly_all_dates, None, "물리대카테", cert_only=True).rename(columns={"tot": "cert_tot", "net": "cert_net"})
        cat_all_ly = A.group_amt_table(ly_amt_data["amt"], ly_all_dates, None, "물리대카테", cert_only=False).rename(columns={"tot": "all_tot", "net": "all_net"})
        cat_ly = pd.merge(cat_cert_ly, cat_all_ly, on="물리대카테", how="outer").fillna(0).set_index("물리대카테")

    sub(f"전체 카테고리 (당월인증거래액 내림차순 · {len(cat)}개)")
    rows = []
    for _, r in cat.iterrows():
        p = cat_ly.loc[r["물리대카테"]] if (cat_ly is not None and r["물리대카테"] in cat_ly.index) else None
        cert_c, all_c = r[f"cert_{ptc}"], r[f"all_{ptc}"]
        cert_p = p[f"cert_{ptc}"] if p is not None else None
        all_p = p[f"all_{ptc}"] if p is not None else None
        rows.append([r["물리대카테"], fmt(cert_c), pct_ratio(cert_c, cert_p)[0], fmt(all_c), pct_ratio(all_c, all_p)[0]])
    df_cat = pd.DataFrame(rows, columns=["카테고리", "당월인증거래액", "당월인증거래액 전년비", "전체거래액", "전체거래액 전년비"])
    st.dataframe(style_delta_cols(df_cat, ["당월인증거래액 전년비", "전체거래액 전년비"]), hide_index=True, use_container_width=True)
    st.caption("전년도 원본 값은 엑셀 다운로드에 포함됩니다.")

    excel_bytes = to_excel_bytes({"카테고리별": cat.assign(**({} if cat_ly is None else {
        "전년_cert_tot": cat["물리대카테"].map(cat_ly["cert_tot"]),
        "전년_cert_net": cat["물리대카테"].map(cat_ly["cert_net"]),
        "전년_all_tot": cat["물리대카테"].map(cat_ly["all_tot"]),
        "전년_all_net": cat["물리대카테"].map(cat_ly["all_net"]),
    }))})
    st.download_button("⤓ 엑셀 다운로드 (카테고리별, 전년도 원본 포함)", excel_bytes, file_name="카테고리별_실적.xlsx")

# ====================================================================================
# 브랜드별 실적 탭
# ====================================================================================
with tab_brand:
    paytype_b = st.radio("결제 구분", ["순결제", "총결제"], horizontal=True, key="brand_paytype")
    ptb = "net" if paytype_b == "순결제" else "tot"

    br_cert = A.group_amt_table(cur_data["amt"], sel_dates, None, "Admin브랜드명", cert_only=True).rename(columns={"tot": "cert_tot", "net": "cert_net"})
    br_all = A.group_amt_table(cur_data["amt"], sel_dates, None, "Admin브랜드명", cert_only=False).rename(columns={"tot": "all_tot", "net": "all_net"})
    br = pd.merge(br_cert, br_all, on="Admin브랜드명", how="outer").fillna(0)
    br = br.sort_values("cert_tot", ascending=False).head(25).reset_index(drop=True)

    br_ly = None
    if not ly_amt_data["amt"].empty:
        br_cert_ly = A.group_amt_table(ly_amt_data["amt"], ly_all_dates, None, "Admin브랜드명", cert_only=True).rename(columns={"tot": "cert_tot", "net": "cert_net"})
        br_all_ly = A.group_amt_table(ly_amt_data["amt"], ly_all_dates, None, "Admin브랜드명", cert_only=False).rename(columns={"tot": "all_tot", "net": "all_net"})
        br_ly = pd.merge(br_cert_ly, br_all_ly, on="Admin브랜드명", how="outer").fillna(0).set_index("Admin브랜드명")

    sub("브랜드 TOP 25 (당월인증거래액 내림차순)")
    rows = []
    for _, r in br.iterrows():
        p = br_ly.loc[r["Admin브랜드명"]] if (br_ly is not None and r["Admin브랜드명"] in br_ly.index) else None
        cert_c, all_c = r[f"cert_{ptb}"], r[f"all_{ptb}"]
        cert_p = p[f"cert_{ptb}"] if p is not None else None
        all_p = p[f"all_{ptb}"] if p is not None else None
        rows.append([r["Admin브랜드명"], fmt(cert_c), pct_ratio(cert_c, cert_p)[0], fmt(all_c), pct_ratio(all_c, all_p)[0]])
    df_br = pd.DataFrame(rows, columns=["브랜드", "당월인증거래액", "당월인증거래액 전년비", "전체거래액", "전체거래액 전년비"])
    st.dataframe(style_delta_cols(df_br, ["당월인증거래액 전년비", "전체거래액 전년비"]), hide_index=True, use_container_width=True)
    st.caption("전년도 원본 값은 엑셀 다운로드에 포함됩니다.")

    excel_bytes = to_excel_bytes({"브랜드별_TOP25": br.assign(**({} if br_ly is None else {
        "전년_cert_tot": br["Admin브랜드명"].map(br_ly["cert_tot"]),
        "전년_cert_net": br["Admin브랜드명"].map(br_ly["cert_net"]),
        "전년_all_tot": br["Admin브랜드명"].map(br_ly["all_tot"]),
        "전년_all_net": br["Admin브랜드명"].map(br_ly["all_net"]),
    }))})
    st.download_button("⤓ 엑셀 다운로드 (브랜드별, 전년도 원본 포함)", excel_bytes, file_name="브랜드별_실적.xlsx")

# ----------------------------------------------------------------------------------
# 임원보고용 종합 리포트 (사이드바 상단 - 항상 접근 가능하도록)
# ----------------------------------------------------------------------------------
st.sidebar.divider()
st.sidebar.header("📑 Step 3. 종합 리포트")
exec_report_bytes = build_exec_report_excel(
    sel_dates, ly_all_dates, yoy_mode_label,
    cur_all, prev_all, cur_cert, prev_cert,
    cert_cur, cert_prev,
    all_cur_d, all_ly_d, cert_cur_d, cert_ly_d,
    pf_cur, pf_ly, target_df,
)
st.sidebar.download_button(
    "⤓ 임원보고용 종합 리포트 다운로드",
    exec_report_bytes,
    file_name=f"LF몰_일반제휴_종합리포트_{sel_dates[0]:%Y%m%d}~{sel_dates[-1]:%Y%m%d}.xlsx",
)
st.sidebar.caption("지표별 실적 요약 · 일자별(신규/윈백/기존)·전년비·목표달성률·MIX · 제휴사별 실적, 순결제 기준 1개 시트")

# ----------------------------------------------------------------------------------
# 전체 JSON 내보내기 (사이드바 하단)
# ----------------------------------------------------------------------------------
st.sidebar.divider()
if st.sidebar.button("⤓ 현재 화면 JSON 스냅샷 만들기"):
    snapshot = {
        "기간": [str(d) for d in sel_dates],
        "UV(전체 제휴사 기준)": uv_cur,
        "인증자수(전체 제휴사 기준)": cert_cur,
        "제휴사별_실적_탭_리스팅_대상(당월인증거래액>0)": qual_partners,
    }
    import json
    st.sidebar.download_button("JSON 다운로드", json.dumps(snapshot, ensure_ascii=False, indent=2),
                                file_name="lf_affiliate_snapshot.json", mime="application/json")
