"""
LF몰 일반제휴 실적 분석 - 데이터 로딩 & 집계 로직 (Streamlit 비의존 모듈)

업로드 3종 파일(당해년도 실적 / 전년도 인증 실적 / 전년도 거래 실적(분기, 복수))을
시트 "이름"이 아니라 "컬럼 구성"으로 자동 분류해서 로딩한다 (파일명/시트명이 달라져도 안전).

[중요] "당월인증거래액>0인 제휴사(qualifying_partners)" 는 제휴사별 실적 탭에 어떤 제휴사를
리스팅할지 정하는 기준일 뿐이다. UV·인증자수·거래액·고객수·객단가 등 개요/일자별/카테고리별/
브랜드별 탭의 모든 집계는 이 모집단으로 제한하지 않고 업로드된 전체 제휴사 기준으로 계산한다
(partners=None 이면 무제한, 명시적으로 리스트를 넘기면 그 제휴사로만 필터링).
"""
from __future__ import annotations
import io
import datetime as dt

import pandas as pd
import numpy as np

try:
    from python_calamine import CalamineWorkbook
    _HAVE_CALAMINE = True
except Exception:  # pragma: no cover - calamine 미설치 환경 대비
    import openpyxl
    _HAVE_CALAMINE = False

STATUSES = ["신규", "WIN-BACK", "기존"]  # 표기: 신규 / 윈백 / 기존


# --------------------------------------------------------------------------------------
# 1) 시트 자동 분류 & 로딩
#    인증거래액류 시트는 5만~15만 행까지 커질 수 있어 python-calamine(러스트 기반, openpyxl 대비
#    5~8배 빠름)을 우선 사용하고, 설치가 안 된 환경에서는 openpyxl read_only로 대체한다.
# --------------------------------------------------------------------------------------
def _to_date(series: pd.Series) -> pd.Series:
    """정산일시일/인증일시일/★일자일 컬럼은 파일마다 datetime 또는 문자열로 섞여 들어옴 -> date로 통일."""
    return pd.to_datetime(series.astype(str).str.slice(0, 10), errors="coerce")


def _read_all_sheets(file_bytes: bytes) -> dict:
    """워크북의 모든 시트를 {시트명: [[row], ...]} (첫 행=헤더) 형태로 읽어온다."""
    if _HAVE_CALAMINE:
        wb = CalamineWorkbook.from_filelike(io.BytesIO(file_bytes))
        return {sn: wb.get_sheet_by_name(sn).to_python() for sn in wb.sheet_names}
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    return {sn: [list(r) for r in wb[sn].iter_rows(values_only=True)] for sn in wb.sheetnames}


def classify_workbook(file_bytes: bytes) -> dict:
    """워크북 안의 시트를 컬럼 구성으로 분류해서 amt/uv/cert/pivot 프레임으로 반환."""
    sheets = _read_all_sheets(file_bytes)
    amt_frames, uv_frames, cert_frames = [], [], []
    pivot_df = None
    for sn, data in sheets.items():
        if not data:
            continue
        if "피벗" in sn:
            try:
                pivot_df = pd.DataFrame(data)
            except Exception:
                pass
            continue
        header = data[0]
        cols = set(map(str, header))
        ncol = len(header)
        body = [r[:ncol] for r in data[1:] if r and r[0] is not None]
        if not body:
            continue
        df = pd.DataFrame(body, columns=header)
        if "거래액_VAT제외" in cols and "정산구분" in cols:
            amt_frames.append(df)
        elif "★UV" in cols:
            uv_frames.append(df)
        elif "총합계" in cols and ("제휴처구분1" in cols or "제휴처구분3" in cols):
            cert_frames.append(df)
    amt = pd.concat(amt_frames, ignore_index=True) if amt_frames else pd.DataFrame()
    uv = pd.concat(uv_frames, ignore_index=True) if uv_frames else pd.DataFrame()
    cert = pd.concat(cert_frames, ignore_index=True) if cert_frames else pd.DataFrame()

    if not amt.empty:
        amt["정산일시일"] = _to_date(amt["정산일시일"])
        amt["거래액_VAT제외"] = pd.to_numeric(amt["거래액_VAT제외"], errors="coerce").fillna(0)
        amt["제휴사"] = amt["제휴사"].astype(str)
        amt.loc[amt["제휴사"].isin(["nan", "None", ""]), "제휴사"] = np.nan
    if not uv.empty:
        uv["★일자일"] = _to_date(uv["★일자일"])
        uv["★UV"] = pd.to_numeric(uv["★UV"], errors="coerce").fillna(0)
    if not cert.empty:
        cert["인증일시일"] = _to_date(cert["인증일시일"])
        for c in ["총합계", "기존", "WIN-BACK", "신규"]:
            if c in cert.columns:
                cert[c] = pd.to_numeric(cert[c], errors="coerce").fillna(0)

    return {"amt": amt, "uv": uv, "cert": cert, "pivot": pivot_df}


def merge_workbooks(list_of_bytes: list) -> dict:
    """전년도 거래 실적처럼 분기별로 여러 파일이 올라오는 경우 합쳐서 하나의 dict로 반환."""
    amt_all, uv_all, cert_all, pivot = [], [], [], None
    for b in list_of_bytes:
        d = classify_workbook(b)
        if not d["amt"].empty:
            amt_all.append(d["amt"])
        if not d["uv"].empty:
            uv_all.append(d["uv"])
        if not d["cert"].empty:
            cert_all.append(d["cert"])
        if pivot is None and d["pivot"] is not None:
            pivot = d["pivot"]
    return {
        "amt": pd.concat(amt_all, ignore_index=True) if amt_all else pd.DataFrame(),
        "uv": pd.concat(uv_all, ignore_index=True) if uv_all else pd.DataFrame(),
        "cert": pd.concat(cert_all, ignore_index=True) if cert_all else pd.DataFrame(),
        "pivot": pivot,
    }


# --------------------------------------------------------------------------------------
# 2) 날짜 유틸: 전년 동일자 매핑 (캘린더 동일 날짜 / 요일기준 52주 전)
# --------------------------------------------------------------------------------------
def ly_date(d: dt.date, mode: str = "calendar") -> dt.date:
    if mode == "calendar":
        try:
            return d.replace(year=d.year - 1)
        except ValueError:  # 2/29 처리
            return d.replace(year=d.year - 1, day=28)
    return d - dt.timedelta(weeks=52)  # weekday 모드: 정확히 52주 전 -> 요일 동일


def ly_dates(dates: list, mode: str = "calendar") -> list:
    return [ly_date(d, mode) for d in dates]


# --------------------------------------------------------------------------------------
# 3) 모집단(scope): 선택 기간에 당월인증거래액이 발생한 제휴사만
# --------------------------------------------------------------------------------------
def qualifying_partners(amt_df: pd.DataFrame, dates: list) -> list:
    if amt_df.empty:
        return []
    sub = amt_df[amt_df["정산일시일"].dt.date.isin(dates) & (amt_df["당월인증"] == "Y")]
    names = [p for p in sub["제휴사"].dropna().unique().tolist() if p and p != "nan"]
    return sorted(names)


# --------------------------------------------------------------------------------------
# 4) 집계 함수
# --------------------------------------------------------------------------------------
def _filter_amt(amt_df, dates, partners, cert_only=False):
    """partners=None 이면 제휴사 제한 없이 전체 제휴사 기준으로 집계한다.
    (참고: '당월인증거래액>0인 제휴사' 모집단 제한은 제휴사별 실적 탭의 리스팅
    기준일 뿐이며, UV/인증자수/거래액 등 다른 모든 집계는 전체 제휴사 기준이다.)"""
    if amt_df.empty:
        return amt_df
    mask = amt_df["정산일시일"].dt.date.isin(dates)
    if partners is not None:
        mask &= amt_df["제휴사"].isin(partners)
    sub = amt_df[mask]
    if cert_only:
        sub = sub[sub["당월인증"] == "Y"]
    return sub


def amt_summary(amt_df, dates, partners=None, cert_only=False) -> dict:
    """구분(전체/신규/윈백/기존)별 총결제·순결제 거래액/고객수. partners=None 이면 전체 제휴사 기준."""
    sub = _filter_amt(amt_df, dates, partners, cert_only)
    statuses = ["전체"] + STATUSES
    if sub.empty:
        return {s: {"tot": 0, "net": 0, "cust_tot": 0, "cust_net": 0} for s in statuses}
    out = {}
    for s in statuses:
        g = sub if s == "전체" else sub[sub["기존/win-back/신규"] == s]
        sale = g[g["정산구분"] == "판매"]
        out[s] = {
            "tot": float(sale["거래액_VAT제외"].sum()),
            "net": float(g["거래액_VAT제외"].sum()),
            "cust_tot": int(sale["고객번호"].nunique()),
            "cust_net": int(g["고객번호"].nunique()),
        }
    return out


def uv_total(uv_df, dates, partners=None, exclude_af=None, name_col="제휴사명") -> int:
    """partners=None 이면 전체 제휴사 기준."""
    if uv_df.empty:
        return 0
    mask = uv_df["★일자일"].dt.date.isin(dates)
    if partners is not None:
        mask &= uv_df[name_col].isin(partners)
    sub = uv_df[mask]
    if exclude_af:
        sub = sub[sub["AF코드"] != exclude_af]
    return int(sub["★UV"].sum())


def cert_summary(cert_df, dates, partners=None, name_col="제휴사") -> dict:
    """partners=None 이면 전체 제휴사 기준."""
    if cert_df.empty:
        return {"전체": 0, "기존": 0, "WIN-BACK": 0, "신규": 0}
    mask = cert_df["인증일시일"].dt.date.isin(dates)
    if partners is not None:
        mask &= cert_df[name_col].isin(partners)
    sub = cert_df[mask]
    return {
        "전체": int(sub["총합계"].sum()),
        "기존": int(sub["기존"].sum()),
        "WIN-BACK": int(sub["WIN-BACK"].sum()),
        "신규": int(sub["신규"].sum()),
    }


def daily_amt_table(amt_df, dates, partners=None, cert_only=False) -> pd.DataFrame:
    """일자 x 구분(전체/신규/윈백/기존) 별 총결제/순결제 거래액·고객수 - 일자별 탭용.
    partners=None 이면 전체 제휴사 기준."""
    sub = _filter_amt(amt_df, dates, partners, cert_only)
    dates_sorted = sorted(dates)
    if sub.empty:
        return pd.DataFrame({"일자": dates_sorted})
    sale = sub[sub["정산구분"] == "판매"]
    g_tot = sale.groupby(["정산일시일", "기존/win-back/신규"])["거래액_VAT제외"].sum()
    g_net = sub.groupby(["정산일시일", "기존/win-back/신규"])["거래액_VAT제외"].sum()
    all_tot = sale.groupby("정산일시일")["거래액_VAT제외"].sum()
    all_net = sub.groupby("정산일시일")["거래액_VAT제외"].sum()
    allc_tot = sale.groupby("정산일시일")["고객번호"].nunique()
    allc_net = sub.groupby("정산일시일")["고객번호"].nunique()
    rows = []
    for d in dates_sorted:
        ts = pd.Timestamp(d)
        row = {
            "일자": d,
            "전체_tot": float(all_tot.get(ts, 0)), "전체_net": float(all_net.get(ts, 0)),
            "전체_cust_tot": int(allc_tot.get(ts, 0)), "전체_cust_net": int(allc_net.get(ts, 0)),
        }
        for s in STATUSES:
            row[f"{s}_tot"] = float(g_tot.get((ts, s), 0))
            row[f"{s}_net"] = float(g_net.get((ts, s), 0))
        rows.append(row)
    return pd.DataFrame(rows)


def daily_uv_cert_table(uv_df, cert_df, dates, partners=None, exclude_af=None) -> pd.DataFrame:
    """partners=None 이면 전체 제휴사 기준."""
    dates_sorted = sorted(dates)
    uv_by_day = {}
    if not uv_df.empty:
        mask = uv_df["★일자일"].dt.date.isin(dates)
        if partners is not None:
            mask &= uv_df["제휴사명"].isin(partners)
        sub = uv_df[mask]
        if exclude_af:
            sub = sub[sub["AF코드"] != exclude_af]
        uv_by_day = sub.groupby(sub["★일자일"].dt.date)["★UV"].sum().to_dict()
    cert_g = {}
    if not cert_df.empty:
        mask = cert_df["인증일시일"].dt.date.isin(dates)
        if partners is not None:
            mask &= cert_df["제휴사"].isin(partners)
        sub = cert_df[mask]
        for col in ["총합계", "기존", "WIN-BACK", "신규"]:
            cert_g[col] = sub.groupby(sub["인증일시일"].dt.date)[col].sum().to_dict()
    rows = []
    for d in dates_sorted:
        rows.append({
            "일자": d,
            "UV": int(uv_by_day.get(d, 0)),
            "인증_전체": int(cert_g.get("총합계", {}).get(d, 0)),
            "인증_신규": int(cert_g.get("신규", {}).get(d, 0)),
            "인증_윈백": int(cert_g.get("WIN-BACK", {}).get(d, 0)),
            "인증_기존": int(cert_g.get("기존", {}).get(d, 0)),
        })
    return pd.DataFrame(rows)


def group_amt_table(amt_df, dates, partners, groupcol, cert_only=False) -> pd.DataFrame:
    """카테고리/브랜드/제휴사 등 임의 그룹컬럼 기준 총결제/순결제 거래액 집계.
    partners=None 이면 전체 제휴사 기준."""
    sub = _filter_amt(amt_df, dates, partners, cert_only)
    if sub.empty:
        return pd.DataFrame(columns=[groupcol, "tot", "net"])
    sale = sub[sub["정산구분"] == "판매"]
    tot = sale.groupby(groupcol)["거래액_VAT제외"].sum()
    net = sub.groupby(groupcol)["거래액_VAT제외"].sum()
    return pd.DataFrame({"tot": tot, "net": net}).reset_index()


def partner_full_table(amt_df, uv_df, cert_df, dates, partners=None, exclude_af=None) -> pd.DataFrame:
    """제휴사별 실적 탭 - UV/인증수/당월인증거래액/전체거래액/고객수/객단가 원본 테이블.
    partners에 qualifying_partners() 결과(당월인증거래액>0인 제휴사)를 넘기면 그 제휴사만
    리스팅한다(제휴사별 실적 탭의 기본 사용법). partners=None 이면 기간 내 활동이 있었던
    전체 제휴사를 리스팅한다(전년도 비교값 조회 등에 사용)."""
    cert_amt = group_amt_table(amt_df, dates, partners, "제휴사", cert_only=True).rename(
        columns={"tot": "cert_tot", "net": "cert_net"})
    all_amt = group_amt_table(amt_df, dates, partners, "제휴사", cert_only=False).rename(
        columns={"tot": "all_tot", "net": "all_net"})
    merged = pd.merge(cert_amt, all_amt, on="제휴사", how="outer").fillna(0)

    sub = _filter_amt(amt_df, dates, partners, cert_only=False)
    if not sub.empty:
        sale = sub[sub["정산구분"] == "판매"]
        cust_tot = sale.groupby("제휴사")["고객번호"].nunique()
        cust_net = sub.groupby("제휴사")["고객번호"].nunique()
        merged["cust_tot"] = merged["제휴사"].map(cust_tot).fillna(0).astype(int)
        merged["cust_net"] = merged["제휴사"].map(cust_net).fillna(0).astype(int)
    else:
        merged["cust_tot"] = 0
        merged["cust_net"] = 0

    effective_partners = partners if partners is not None else merged["제휴사"].tolist()
    uv_map, cert_map = {}, {}
    for p in effective_partners:
        uv_map[p] = uv_total(uv_df, dates, [p], exclude_af=exclude_af)
        cert_map[p] = cert_summary(cert_df, dates, [p])["전체"]
    merged["UV"] = merged["제휴사"].map(uv_map)
    merged["인증수"] = merged["제휴사"].map(cert_map)
    return merged.sort_values("cert_tot", ascending=False).reset_index(drop=True)


def partner_daily_table(amt_df, uv_df, cert_df, dates, partner, exclude_af=None) -> pd.DataFrame:
    d = daily_amt_table(amt_df, dates, [partner], cert_only=False)
    dc = daily_amt_table(amt_df, dates, [partner], cert_only=True)
    uc = daily_uv_cert_table(uv_df, cert_df, dates, [partner], exclude_af=exclude_af)
    out = uc.merge(d[["일자", "전체_tot", "전체_net", "전체_cust_tot", "전체_cust_net"]], on="일자", how="left")
    out = out.merge(dc[["일자", "전체_tot", "전체_net"]].rename(
        columns={"전체_tot": "당월인증_tot", "전체_net": "당월인증_net"}), on="일자", how="left")
    return out.rename(columns={"전체_tot": "전체거래액_tot", "전체_net": "전체거래액_net",
                                "전체_cust_tot": "고객수_tot", "전체_cust_net": "고객수_net"})


# --------------------------------------------------------------------------------------
# 5) 일자별 목표(순결제 기준) 로딩 & 조회 - "목표비" 기능용
#    파일 구조: 3행 병합헤더('구분'/'목표 거래액' -> '전체 인증 기준'/'당월인증 기준' ->
#    'total'/'신규'/'윈백'), 4행부터 일자별 데이터. 기존 = total - 신규 - 윈백로 계산한다.
# --------------------------------------------------------------------------------------
def load_target_file(file_bytes: bytes) -> pd.DataFrame:
    """반환 컬럼: 일자, all_total/all_신규/all_윈백/all_기존, cert_total/cert_신규/cert_윈백/cert_기존
    (all_ = "전체 인증 기준" = 전체거래액 목표, cert_ = "당월인증 기준" = 당월인증거래액 목표,
    전부 순결제 기준. 기존 = total - 신규 - 윈백로 계산.)"""
    sheets = _read_all_sheets(file_bytes)
    target_sheet = None
    for sn, data in sheets.items():
        if "목표" in sn:
            target_sheet = data
            break
    if target_sheet is None and sheets:
        target_sheet = next(iter(sheets.values()))
    if not target_sheet:
        return pd.DataFrame()

    header_rows = target_sheet[:5]
    all_start = cert_start = None
    for row in header_rows:
        for j, c in enumerate(row):
            s = str(c)
            if all_start is None and "전체" in s and "인증" in s:
                all_start = j
            if cert_start is None and "당월인증" in s:
                cert_start = j
    if all_start is None or cert_start is None:
        all_start, cert_start = 6, 9  # 관찰된 파일 구조로 폴백

    def _offsets(start):
        new_off, wb_off = 1, 2
        for row in header_rows:
            for j in range(start, min(start + 3, len(row))):
                s = str(row[j])
                if "신규" in s:
                    new_off = j - start
                if "윈백" in s or "WIN" in s.upper():
                    wb_off = j - start
        return new_off, wb_off

    all_new_off, all_wb_off = _offsets(all_start)
    cert_new_off, cert_wb_off = _offsets(cert_start)

    start_row = None
    for i, row in enumerate(target_sheet):
        if row and isinstance(row[0], (dt.date, dt.datetime)):
            start_row = i
            break
    if start_row is None:
        return pd.DataFrame()

    def _num(row, idx):
        try:
            v = row[idx]
            return float(v) if v not in (None, "") else 0.0
        except (IndexError, TypeError, ValueError):
            return 0.0

    rows = []
    for row in target_sheet[start_row:]:
        if not row or row[0] is None:
            continue
        d = row[0]
        if isinstance(d, dt.datetime):
            d = d.date()
        if not isinstance(d, dt.date):
            continue
        all_total = _num(row, all_start)
        all_new = _num(row, all_start + all_new_off)
        all_wb = _num(row, all_start + all_wb_off)
        cert_total = _num(row, cert_start)
        cert_new = _num(row, cert_start + cert_new_off)
        cert_wb = _num(row, cert_start + cert_wb_off)
        rows.append({
            "일자": d,
            "all_total": all_total, "all_신규": all_new, "all_윈백": all_wb,
            "all_기존": all_total - all_new - all_wb,
            "cert_total": cert_total, "cert_신규": cert_new, "cert_윈백": cert_wb,
            "cert_기존": cert_total - cert_new - cert_wb,
        })
    return pd.DataFrame(rows)


def target_sum(target_df: pd.DataFrame, dates: list, group: str, status: str = "total"):
    """선택 기간(dates)에 해당하는 일별 목표 합계. 매칭되는 날짜가 하나도 없으면 None."""
    if target_df is None or target_df.empty:
        return None
    col = f"{group}_{status}"
    if col not in target_df.columns:
        return None
    sub = target_df[target_df["일자"].isin(dates)]
    if sub.empty:
        return None
    return float(sub[col].sum())


def target_lookup(target_df: pd.DataFrame, date, group: str, status: str = "total"):
    """특정 하루의 목표값. 목표 파일에 해당 날짜가 없으면 None."""
    if target_df is None or target_df.empty:
        return None
    col = f"{group}_{status}"
    if col not in target_df.columns:
        return None
    sub = target_df[target_df["일자"] == date]
    if sub.empty:
        return None
    return float(sub.iloc[0][col])
